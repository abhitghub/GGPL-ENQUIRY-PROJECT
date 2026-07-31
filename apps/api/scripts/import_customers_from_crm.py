"""Import a CRM contact export (xlsx) into the customer master.

Each row is a contact; contacts are grouped by the ``Account Name`` column into
customers, each holding a ``contacts`` list. Rows without an Account Name are
skipped. Writes into the local API datastore under
``app_settings['<org>:customers']`` (preserving epc_names). Restart the API
afterwards so it reloads the datastore.

Every row also carries its own address, and one company routinely spans several
plants (L&T alone has a dozen). Those distinct addresses become the company's
``locations`` list and each contact is pinned to theirs via ``location_id``, so
picking a contact person in the enquiry setup fills the right site address and a
multi-site company offers the sites as a dropdown.

Usage:
    python -m scripts.import_customers_from_crm <export.xlsx> [--org local-org] \
        [--store ../../.local/api_repository.json] [--sheet "Master Customers"]
"""

from __future__ import annotations

import argparse
import json
import re
from collections import OrderedDict
from pathlib import Path

import openpyxl


def _clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def _norm(value: str) -> str:
    """Fold case, spacing and trailing punctuation so 'HAZIRA' and 'Hazira,' are
    recognised as the same site instead of becoming two dropdown entries."""
    return re.sub(r"\s+", " ", value).strip(" ,.|").upper()


PLACE_FIELDS = ("address_line1", "city", "state", "pin_code", "country")


def _location_key(loc: dict) -> tuple:
    """What makes two rows the same plant: street, city and state.

    The PIN and country are attributes of a site rather than what distinguishes
    one from another — keying on them too would list 'HAZIRA, Gujarat' twice
    because one row left the country blank.
    """
    return tuple(_norm(loc[k]) for k in ("address_line1", "city", "state"))


def _absorb(into: dict, extra: dict) -> None:
    """Fill the gaps of a site from a less complete record of the same site."""
    for field in (*PLACE_FIELDS, "region"):
        if not into[field] and extra[field]:
            into[field] = extra[field]


def _merge_partial_locations(locations: list[dict]) -> tuple[list[dict], dict[str, str]]:
    """Fold sites recorded with gaps into the fuller record of the same site.

    A row that names only 'Gujarat' where another names 'HAZIRA, Gujarat, 394510'
    is the same plant entered by someone in a hurry, not a second one. Such a row
    is absorbed only when exactly one fuller site could be meant — two candidates
    means the true site is genuinely unknown, and inventing one would put the
    wrong address on a quotation.

    Returns the surviving sites and a map from every dropped id to its survivor,
    so the contacts pinned to a dropped id can be repointed.
    """
    def filled(loc: dict) -> set:
        return {field for field in PLACE_FIELDS if _norm(loc[field])}

    # Fuller records first, so a gappy row is always tested against a superset.
    order = sorted(locations, key=lambda loc: len(filled(loc)), reverse=True)
    survivors: list[dict] = []
    remap: dict[str, str] = {}
    for loc in order:
        keys = filled(loc)
        candidates = [
            other
            for other in survivors
            if keys < filled(other) and all(_norm(loc[field]) == _norm(other[field]) for field in keys)
        ]
        if len(candidates) == 1:
            _absorb(candidates[0], loc)
            remap[loc["id"]] = candidates[0]["id"]
            continue
        survivors.append(loc)
    # Keep the original workbook order of whatever survived.
    kept = {loc["id"] for loc in survivors}
    return [loc for loc in locations if loc["id"] in kept], remap


def _location_label(loc: dict) -> str:
    parts = [loc["city"], loc["state"], loc["country"]]
    seen: list[str] = []
    for part in parts:
        part = re.sub(r"\s+", " ", part).strip(" ,.|")
        # City and state are often the same word in this export ("ODISHA",
        # "Odisha") — printing it twice reads like a data error in the dropdown.
        if part and _norm(part) not in {_norm(x) for x in seen}:
            seen.append(part)
    return ", ".join(seen)


def build_customers(xlsx: Path, sheet: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    header = [_clean(c) for c in next(rows)]
    idx = {h: i for i, h in enumerate(header)}

    def g(row: tuple, name: str) -> str:
        i = idx.get(name)
        return "" if i is None or i >= len(row) else _clean(row[i])

    customers: "OrderedDict[str, dict]" = OrderedDict()
    seen: dict[str, set] = {}
    loc_index: dict[str, dict[tuple, str]] = {}
    seq = 0
    for row in rows:
        account = g(row, "Account Name")
        if not account:
            continue
        key = re.sub(r"\s+", " ", account).strip().upper()
        if key not in customers:
            seq += 1
            customers[key] = {
                "id": f"cust-{seq}",
                "name": re.sub(r"\s+", " ", account).strip(),
                "address_line1": "", "address_line2": "", "city": "", "state": "",
                "pin_code": "", "country": "",
                "contact_name": "", "designation": "", "email": "", "phone": "",
                "gst_no": "", "default_currency": "INR", "payment_terms": "",
                "delivery_terms": "", "active": True, "contacts": [], "locations": [],
            }
            seen[key] = set()
            loc_index[key] = {}
        rec = customers[key]

        # This row's own site. Company City/State win over the Mailing pair: the
        # export fills them far more often, and the Mailing columns are shifted
        # by one on a chunk of rows.
        loc = {
            "address_line1": g(row, "Mailing Street"),
            "address_line2": "",
            "city": g(row, "Company City") or g(row, "Mailing City"),
            "state": g(row, "Company State") or g(row, "Mailing State"),
            "pin_code": g(row, "Mailing Zip"),
            "country": g(row, "Company Country") or g(row, "Mailing Country"),
            "region": g(row, "Region"),
            "gst_no": "",
        }
        location_id = ""
        if any(loc[k] for k in PLACE_FIELDS):
            lkey = _location_key(loc)
            location_id = loc_index[key].get(lkey, "")
            if location_id:
                # Same plant seen again — take whatever this row fills in that the
                # first sighting left blank (often the PIN or the country).
                _absorb(next(row for row in rec["locations"] if row["id"] == location_id), loc)
            else:
                location_id = f"{rec['id']}-l{len(rec['locations']) + 1}"
                loc_index[key][lkey] = location_id
                rec["locations"].append({"id": location_id, **loc})

        name = g(row, "Contact Name") or " ".join(x for x in (g(row, "First Name"), g(row, "Last Name")) if x)
        email = g(row, "Email")
        if not name and not email:
            continue
        dedupe = (name.lower(), email.lower())
        if dedupe in seen[key]:
            continue
        seen[key].add(dedupe)
        contact = {
            "id": f"{rec['id']}-c{len(rec['contacts']) + 1}",
            "name": name, "designation": g(row, "Title"), "department": g(row, "Department"),
            "email": email, "phone": g(row, "Phone"), "mobile": g(row, "Mobile"),
            "location_id": location_id,
        }
        rec["contacts"].append(contact)
        if not rec["contact_name"] and (name or email):
            rec.update({"contact_name": name, "designation": contact["designation"],
                        "email": email, "phone": contact["phone"] or contact["mobile"]})

    for rec in customers.values():
        # Fold the gappy duplicates together before labelling, so a label is
        # built from the fullest version of each site.
        rec["locations"], remap = _merge_partial_locations(rec["locations"])
        for contact in rec["contacts"]:
            contact["location_id"] = remap.get(contact["location_id"], contact["location_id"])
        for loc in rec["locations"]:
            loc["label"] = _location_label(loc)
        # The company's own address stays as the fallback for contacts with no
        # site of their own, and for the consumers that read a single address.
        if rec["locations"]:
            first = rec["locations"][0]
            for field in ("address_line1", "address_line2", *PLACE_FIELDS[1:]):
                rec[field] = first[field]
    return list(customers.values())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", type=Path)
    ap.add_argument("--org", default="local-org")
    ap.add_argument("--sheet", default="Master Customers")
    ap.add_argument("--store", type=Path, default=Path(__file__).resolve().parents[1] / ".local" / "api_repository.json")
    args = ap.parse_args()

    customers = build_customers(args.xlsx, args.sheet)
    total_contacts = sum(len(c["contacts"]) for c in customers)
    print(f"companies: {len(customers)}  contacts: {total_contacts}")

    store = json.loads(args.store.read_text(encoding="utf-8")) if args.store.exists() else {}
    settings = store.setdefault("app_settings", {})
    key = f"{args.org}:customers"
    epc_names = (settings.get(key) or {}).get("epc_names") or []
    settings[key] = {"customers": customers, "epc_names": epc_names}
    args.store.write_text(json.dumps(store, ensure_ascii=False, indent=2), encoding="utf-8")
    print("written:", args.store)


if __name__ == "__main__":
    main()
