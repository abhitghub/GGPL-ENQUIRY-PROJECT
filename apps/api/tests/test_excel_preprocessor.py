from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook

from packages.core.document_reader import _csv_to_text, _excel_to_text


def _save_workbook(wb: Workbook) -> bytes:
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def test_excel_to_text_prefers_enquiry_check_sheet_in_review_workbook():
    wb = Workbook()
    summary = wb.active
    summary.title = "1-Summary Dashboard"
    summary.append(["Metric", "Value"])
    summary.append(["Total Line Items", 2])

    enquiry = wb.create_sheet("2-Enquiry Check")
    enquiry.append([
        "S.No",
        "Customer Enquiry Description",
        "Detected Product Type",
        "GGPL Quote Description",
        "Qty",
        "UoM",
    ])
    enquiry.append([1, "6 INCH CLASS 300 INSULATION KIT", "ISK", "6 x 300# ISK", 4, "KIT"])
    enquiry.append([2, "2 INCH CLASS 150 RTJ GASKET", "RTJ", "R-23 RTJ SS316", 10, "NOS"])

    deviations = wb.create_sheet("3-Deviations Log")
    deviations.append(["S.No", "Issue Type", "Customer Said", "Action Required", "Severity"])
    deviations.append([1, "Missing class", "size only", "Confirm rating", "RED"])

    text, truncated, row_count = _excel_to_text(_save_workbook(wb))

    assert truncated is False
    assert row_count == 2
    assert "=== Sheet: 2-Enquiry Check ===" in text
    assert "=== Sheet: 1-Summary Dashboard ===" not in text
    assert "=== Sheet: 3-Deviations Log ===" not in text
    assert "6 INCH CLASS 300 INSULATION KIT" in text
    assert "Missing class" not in text


def test_excel_to_text_keeps_plain_customer_sheet_without_review_tabs():
    wb = Workbook()
    sheet = wb.active
    sheet.title = "RFQ"
    sheet.append(["Description", "Qty", "UoM"])
    sheet.append(["SPIRAL WOUND GASKET 4 INCH 150#", 8, "NOS"])

    text, truncated, row_count = _excel_to_text(_save_workbook(wb))

    assert truncated is False
    assert row_count == 1
    assert "=== Sheet: RFQ ===" in text
    assert "SPIRAL WOUND GASKET 4 INCH 150#" in text


def test_csv_to_text_repairs_reference_csv_with_description_in_header():
    source = (
        '"GASKET,FLANGE NONSPIRAL;FULL FACE;20"" PIPE;150LB;TEFLON;QA/QC CERT REQ A,C,D",'
        'Customer,Reference,S No,Discription,GGPL DESCRIPTPION,Deviations,TYPE OF PRODUCT,NON STANDARD\n'
        ',L&T,RFQ-1,2,"NPS 2, Gasket Spiral wound, SS316 with flexible graphite filler",'
        '"SIZE : 2"" X 150# X 4.5MM THK ,SS316 SPIRAL WOUND GASKET",,Spiral Wound,\n'
    )

    text, truncated, row_count = _csv_to_text(source)

    assert truncated is False
    assert row_count == 2
    assert "=== Sheet: CSV ===" in text
    assert "Customer Enquiry Description" in text
    assert "| CSV | 1 | 1 |" in text
    assert "| CSV | 2 | 2 |" in text
    assert 'GASKET,FLANGE NONSPIRAL' in text
    assert "NPS 2, Gasket Spiral wound" in text
    assert "GGPL Quote Description" in text


def test_csv_to_text_handles_plain_header_csv():
    source = (
        "Customer Enquiry Description,Qty,UoM\n"
        "\"RTJ Ring number - 23 , Moc :- Inconel 825 ,Hardness required - 160 HBN\",4,NOS\n"
    )

    text, truncated, row_count = _csv_to_text(source.encode("utf-8"))

    assert truncated is False
    assert row_count == 1
    assert "RTJ Ring number - 23" in text
    assert "| CSV | 2 | 1 |" in text
