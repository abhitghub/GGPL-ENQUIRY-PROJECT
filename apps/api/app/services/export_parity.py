from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Any

import openpyxl
import pdfplumber


@dataclass(frozen=True)
class ExcelDiff:
    location: str
    field: str
    expected: Any
    actual: Any


def _style_fingerprint(cell: Any) -> tuple[Any, ...]:
    fill = cell.fill
    font = cell.font
    alignment = cell.alignment
    border = cell.border
    return (
        cell.number_format,
        fill.fill_type,
        fill.fgColor.rgb,
        font.name,
        font.sz,
        font.bold,
        font.italic,
        font.color.rgb if font.color and font.color.type == "rgb" else None,
        alignment.horizontal,
        alignment.vertical,
        border.left.style,
        border.right.style,
        border.top.style,
        border.bottom.style,
    )


def compare_excel(expected: bytes, actual: bytes) -> list[ExcelDiff]:
    expected_book = openpyxl.load_workbook(io.BytesIO(expected), data_only=False)
    actual_book = openpyxl.load_workbook(io.BytesIO(actual), data_only=False)
    diffs: list[ExcelDiff] = []
    if expected_book.sheetnames != actual_book.sheetnames:
        diffs.append(ExcelDiff("workbook", "sheetnames", expected_book.sheetnames, actual_book.sheetnames))
        return diffs
    for sheet_name in expected_book.sheetnames:
        expected_sheet = expected_book[sheet_name]
        actual_sheet = actual_book[sheet_name]
        expected_dims = (expected_sheet.max_row, expected_sheet.max_column)
        actual_dims = (actual_sheet.max_row, actual_sheet.max_column)
        if expected_dims != actual_dims:
            diffs.append(ExcelDiff(sheet_name, "dimensions", expected_dims, actual_dims))
            continue
        for row in range(1, expected_sheet.max_row + 1):
            for col in range(1, expected_sheet.max_column + 1):
                expected_cell = expected_sheet.cell(row, col)
                actual_cell = actual_sheet.cell(row, col)
                location = f"{sheet_name}!{expected_cell.coordinate}"
                if expected_cell.value != actual_cell.value:
                    diffs.append(ExcelDiff(location, "value", expected_cell.value, actual_cell.value))
                if _style_fingerprint(expected_cell) != _style_fingerprint(actual_cell):
                    diffs.append(
                        ExcelDiff(location, "style", _style_fingerprint(expected_cell), _style_fingerprint(actual_cell))
                    )
                if expected_cell.number_format != actual_cell.number_format:
                    diffs.append(ExcelDiff(location, "number_format", expected_cell.number_format, actual_cell.number_format))
    return diffs


def extract_pdf_text(pdf_bytes: bytes) -> str:
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages).strip()


def compare_pdf_text(expected: bytes, actual: bytes) -> tuple[bool, str, str]:
    expected_text = extract_pdf_text(expected)
    actual_text = extract_pdf_text(actual)
    return expected_text == actual_text, expected_text, actual_text
