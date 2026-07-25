"""Evaluate the deterministic conversion pipeline against the
'Customer Data - 3rd july .xlsx' ground-truth workbook (EXPORT + Domestic).

For every unique customer description the script runs the same deterministic
path the app uses (enrich -> rules -> format) and compares the produced GGPL
description against the human-written one (normalised: uppercase,
alphanumeric-only). Results are reported per gasket-type label.

Usage:  python scripts/evaluate_customer_data_workbook.py [path-to-workbook]
"""
from __future__ import annotations

import collections
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
for p in (ROOT, ROOT / "packages"):
    if str(p) not in sys.path:
        sys.path.insert(0, str(p))

from core.parser import _enrich_from_description  # noqa: E402
from core.rules import apply_rules  # noqa: E402
from core.formatter import format_description  # noqa: E402

DEFAULT_WORKBOOK = ROOT / "Customer Data - 3rd july .xlsx"

# (sheet, customer-description column, ggpl-description column, type column)
SHEETS = [
    ("EXPORT", 4, 5, 7),
    ("Domestic", 3, 4, 6),
]


def load_pairs(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    uniq: dict[str, dict] = {}
    for sheet, d_col, g_col, t_col in SHEETS:
        if sheet not in wb.sheetnames:
            continue
        for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
            if i == 0:
                continue
            d = row[d_col] if d_col < len(row) else None
            g = row[g_col] if g_col < len(row) else None
            t = row[t_col] if t_col is not None and t_col < len(row) else None
            if not (d and g):
                continue
            key = " ".join(str(d).upper().split())
            if key not in uniq:
                uniq[key] = {
                    "desc": str(d).strip(),
                    "ggpl": str(g).strip(),
                    "type": str(t).strip() if t else "?",
                }
    return list(uniq.values())


def norm(text: str | None) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(text or "").upper())


def convert(desc: str) -> str:
    item = {"raw_description": desc, "description": desc}
    try:
        item = apply_rules(_enrich_from_description(item))
        return format_description(item) or ""
    except Exception as e:  # noqa: BLE001
        return f"<ERROR: {e}>"


def main() -> int:
    path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    pairs = load_pairs(path)
    by_type: dict[str, list[int]] = collections.defaultdict(lambda: [0, 0])
    mismatches = []
    for rec in pairs:
        actual = convert(rec["desc"])
        ok = norm(actual) == norm(rec["ggpl"])
        by_type[rec["type"]][1] += 1
        if ok:
            by_type[rec["type"]][0] += 1
        else:
            mismatches.append({**rec, "actual": actual})

    total = len(pairs)
    matched = total - len(mismatches)
    print(f"TOTAL: {matched}/{total} exact-normalised matches ({matched / total * 100:.1f}%)")
    print("\nBy labelled type:")
    for t, (m, n) in sorted(by_type.items(), key=lambda kv: -kv[1][1]):
        print(f"  {t:30s} {m:6d}/{n:<6d} ({m / n * 100:5.1f}%)")

    out = ROOT / "scripts" / "customer_data_eval_mismatches.json"
    out.write_text(json.dumps(mismatches, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nmismatches written to {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
