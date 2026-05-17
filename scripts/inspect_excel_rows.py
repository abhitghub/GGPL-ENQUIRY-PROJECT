from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl


def main() -> int:
    path = Path(sys.argv[1])
    max_rows = int(sys.argv[2]) if len(sys.argv) > 2 else 20
    wb = openpyxl.load_workbook(path, data_only=True)
    for ws in wb.worksheets:
        print(f"SHEET {ws.title} rows={ws.max_row} cols={ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row), values_only=True):
            print(json.dumps([str(value) if value is not None else None for value in row], ensure_ascii=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
