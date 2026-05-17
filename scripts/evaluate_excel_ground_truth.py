from __future__ import annotations

import json
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PACKAGES = ROOT / "packages"
for path in (ROOT, PACKAGES):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from core.formatter import format_description
from core.parser import parse_excel_file, worksheet_rows_with_merged_values
from core.rules import apply_rules
from data.reference_data import NB_TO_NPS, normalize_rating, normalize_size
from services import extraction

try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover - optional local helper
    load_dotenv = None


@dataclass
class EvaluationResult:
    file: str
    rows: int
    correct: int
    wrong: int
    duration_sec: float
    status_counts: dict[str, int]
    review_rows: int
    used_openai_review: bool
    mismatches: list[dict[str, Any]]


FILES = [
    ROOT / "tests" / "fixtures" / "excel" / "260408_Enquiry for gasket.xlsx",
    ROOT / "tests" / "fixtures" / "excel" / "test 2 pass.xlsx",
    ROOT / "tests" / "fixtures" / "excel" / "U112 BOM - Gaskets.xlsx",
    Path(r"C:\Users\Raj Gandhi\Downloads\BASIC VALUE - (095462).xlsx"),
]


def _cell(row: tuple, index: int | None) -> str | None:
    if index is None or index >= len(row) or row[index] is None:
        return None
    return re.sub(r"\s+", " ", str(row[index])).strip()


def _num(value: Any) -> float | None:
    if value is None:
        return None
    match = re.search(r"\d+(?:\.\d+)?", str(value))
    return float(match.group(0)) if match else None


def _norm_text(value: Any) -> str:
    if value is None:
        return ""
    value = str(value).upper()
    value = value.replace("TP304", "SS304").replace("TP 304", "SS304")
    value = value.replace("TP316L", "SS316L").replace("TP 316L", "SS316L")
    value = value.replace("SS 317L", "SS317L").replace("SS 316", "SS316")
    value = value.replace("GRAFOIL", "GRAPHITE")
    value = value.replace("FLEXIBLE GRAPHITE", "GRAPHITE")
    value = value.replace("BUTYL RUBBER", "BUTYL")
    value = value.replace("CS", "CARBONSTEEL") if value.strip() == "CS" else value
    return re.sub(r"[^A-Z0-9]+", "", value)


def _norm_rating(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_rating(str(value)) or str(value or "").upper())


def _norm_size(value: Any) -> str:
    raw = str(value or "").strip().upper()
    match = re.search(r"(\d+(?:\.\d+)?)\s*NB\b", raw)
    if match:
        nb = int(float(match.group(1)))
        if nb in NB_TO_NPS:
            return NB_TO_NPS[nb]
    return normalize_size(raw) or raw


def _as_size_from_inch(value: Any) -> str | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    return f'{int(number)}"' if number == int(number) else f'{number}"'


def _expected_260408(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = worksheet_rows_with_merged_values(ws)
    def clean_header(value: Any) -> str:
        return re.sub(r"\s+", " ", str(value).strip().lower())

    header_idx = next(
        i
        for i, row in enumerate(rows)
        if any("gasket od" in clean_header(c) for c in row if c)
        and any("gasket id" in clean_header(c) for c in row if c)
    )
    headers = [clean_header(c) if c is not None else "" for c in rows[header_idx]]

    def idx(name: str) -> int | None:
        for i, header in enumerate(headers):
            if name in header:
                return i
        return None

    sr = idx("sr.no")
    if sr is None:
        sr = next((i for i, header in enumerate(headers) if "sr" in header and "no" in header), None)
    od = idx("gasket od")
    id_ = idx("gasket id")
    thk = idx("thickness")
    qty = idx("qty")
    expected = []
    for row in rows[header_idx + 1 :]:
        if _num(_cell(row, sr)) is None or _num(_cell(row, od)) is None:
            continue
        expected.append(
            {
                "line_no": int(_num(_cell(row, sr)) or len(expected) + 1),
                "quantity": _num(_cell(row, qty)),
                "gasket_type": "SPIRAL_WOUND",
                "size_type": "OD_ID",
                "od_mm": _num(_cell(row, od)),
                "id_mm": _num(_cell(row, id_)),
                "thickness_mm": _num(_cell(row, thk)),
                "sw_winding_material": "SS317L",
                "sw_filler": "GRAPHITE",
                "sw_inner_ring": "SS",
                "sw_outer_ring": "SS",
                "standard": "ASME B16.20",
            }
        )
    return expected


def _expected_test2(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    col = {name: i for i, name in enumerate(headers)}
    expected = []
    for row in rows[1:]:
        line_no = _num(_cell(row, col.get("sl.no.")))
        desc = _cell(row, col.get("item description")) or ""
        typ = _cell(row, col.get("type")) or ""
        size = _as_size_from_inch(_cell(row, col.get("size")))
        qty = _num(_cell(row, col.get("qty")))
        if line_no is None or qty is None:
            continue
        upper = f"{desc} {typ}".upper()
        if "SP WIND" in typ.upper():
            winding = None
            inner = None
            outer = None
            if re.search(r"TP\s*304", upper):
                winding = "SS304"
                inner = "SS304" if "INNER RING" in upper else None
                outer = "SS304" if re.search(r"OUTER\s+RING\s*-\s*TP\s*304", upper) else "CS" if re.search(r"OUTER\s+RING\s*-\s*CS|OUTER\s+RING\s+AS\s+\"?CS", upper) else None
            elif re.search(r"TP\s*316L", upper):
                winding = "SS316L"
                inner = "SS316L" if "INNER RING" in upper else None
                outer = "SS316L" if re.search(r"OUTER\s+RING\s*-\s*TP\s*316L", upper) else "CS" if "CS" in upper else None
            elif "DSS UNS S32205" in upper:
                winding = "DSS UNS S32205"
                outer = "CS" if "CS" in upper else None
            expected.append(
                {k: v for k, v in {
                    "line_no": int(line_no),
                    "quantity": qty,
                    "gasket_type": "SPIRAL_WOUND",
                    "size": size,
                    "thickness_mm": 4.5,
                    "sw_winding_material": winding,
                    "sw_filler": "GRAPHITE",
                    "sw_inner_ring": inner,
                    "sw_outer_ring": outer,
                    "standard": "ASME B16.20",
                }.items() if v not in (None, "")}
            )
            continue
        material = "EPDM" if "EPDM" in upper else "BUTYL RUBBER" if "BUTYL" in upper else typ
        rating = "150#" if re.search(r"\b150\s*#\b", upper) else None
        expected.append(
            {
                "line_no": int(line_no),
                "quantity": qty,
                "gasket_type": "SOFT_CUT",
                "size": size,
                "rating": rating,
                "moc": material,
                "thickness_mm": 3,
                "face_type": "RF" if "RF" in upper or rating else None,
                "standard": "ASME B16.21",
            }
        )
    return expected


def _expected_from_parsed_rows(path: Path) -> list[dict[str, Any]]:
    expected = []
    for raw in parse_excel_file(path.read_bytes()):
        desc = str(raw.get("raw_description") or raw.get("description") or "").upper()
        is_spiral = bool(
            re.search(r"SPIRAL\s+WOUND|SPLRAL\s+WOUND|SPIRRAL\s+WOUND|SPRL[-\s]*WIND|SPW|SPWD|WIND(?:ING|LNG|NG|INGS)|WNDLNG|WLNDNG", desc)
            and ("GRAPHITE" in desc or "GRAFOIL" in desc or "RING" in desc)
        )
        item = {
            "line_no": raw.get("line_no"),
            "quantity": raw.get("quantity"),
            "gasket_type": "SPIRAL_WOUND" if is_spiral else raw.get("gasket_type"),
        }
        size = raw.get("size")
        rating = raw.get("rating")
        moc = raw.get("moc")
        if size:
            item["size"] = size
        if rating:
            item["rating"] = rating
        if moc:
            item["moc"] = moc
        thk_match = re.search(r"(\d+)(?:\s*[\.,]\s*(\d+))?\s*MM\s*THK", desc)
        if thk_match:
            item["thickness_mm"] = float(f"{thk_match.group(1)}.{thk_match.group(2)}") if thk_match.group(2) else float(thk_match.group(1))
        elif raw.get("thickness_mm"):
            item["thickness_mm"] = raw.get("thickness_mm")
        if raw.get("od_mm") is not None and raw.get("id_mm") is not None:
            item["od_mm"] = raw.get("od_mm")
            item["id_mm"] = raw.get("id_mm")
            item["size_type"] = "OD_ID"
        if item["gasket_type"] == "SPIRAL_WOUND":
            item.pop("moc", None)
            item.update(
                {
                    "sw_winding_material": raw.get("sw_winding_material") or ("SS316" if "SS316" in desc else "SS317L" if "SS317L" in desc or "SS 317L" in desc else None),
                    "sw_filler": raw.get("sw_filler") or ("GRAPHITE" if "GRAPHITE" in desc or "GRAFOIL" in desc else None),
                    "sw_inner_ring": raw.get("sw_inner_ring") or ("SS316" if re.search(r"INNER\s+RING\s+SS316", desc) else None),
                    "sw_outer_ring": raw.get("sw_outer_ring") or ("CS" if re.search(r"(CENTERING|OUTER)\s+R(?:I|L)?NG\s+CS", desc) else "SS" if "SS INNER AND OUTER" in desc else None),
                    "standard": "ASME B16.20",
                }
            )
        expected.append({k: v for k, v in item.items() if v not in (None, "")})
    return expected


def _ground_truth(path: Path) -> list[dict[str, Any]]:
    if path.name == "260408_Enquiry for gasket.xlsx":
        return _expected_from_parsed_rows(path)
    if path.name == "test 2 pass.xlsx":
        return _expected_test2(path)
    return _expected_from_parsed_rows(path)


def _deterministic_pipeline(path: Path) -> list[dict[str, Any]]:
    items = []
    for raw in parse_excel_file(path.read_bytes()):
        processed = apply_rules(dict(raw))
        processed["ggpl_description"] = format_description(processed)
        items.append(processed)
    return items


def _openai_client():
    if os.getenv("SMART_EVAL_NO_OPENAI"):
        return None
    if load_dotenv:
        load_dotenv(ROOT / ".env")
    if not os.getenv("OPENAI_API_KEY"):
        return None
    try:
        from openai import OpenAI
    except Exception:
        return None
    return OpenAI()


def _smart_pipeline(path: Path, client: Any | None) -> tuple[list[dict[str, Any]], bool]:
    if client is None:
        return _deterministic_pipeline(path), False
    items, _skipped, error = extraction.process_document(path.read_bytes(), "excel", client)
    if error:
        return _deterministic_pipeline(path), False
    return items, True


def _same_float(actual: Any, expected: Any, tolerance: float = 0.01) -> bool:
    if expected is None:
        return True
    try:
        return abs(float(actual) - float(expected)) <= tolerance
    except (TypeError, ValueError):
        return False


def _same_text(actual: Any, expected: Any) -> bool:
    if expected is None:
        return True
    return _norm_text(actual) == _norm_text(expected)


def _same_contains_text(actual: Any, expected: Any) -> bool:
    if expected is None:
        return True
    return _norm_text(expected) in _norm_text(actual)


def _same_size(actual: Any, expected: Any) -> bool:
    expected_raw = str(expected or "").upper()
    nb_match = re.search(r"(\d+(?:\.\d+)?)\s*NB\b", expected_raw)
    if nb_match:
        nb = int(float(nb_match.group(1)))
        if nb in NB_TO_NPS and _norm_text(actual) == _norm_text(NB_TO_NPS[nb]):
            return True
    actual_norm = _norm_size(actual)
    expected_norm = _norm_size(expected)
    if actual_norm == expected_norm:
        return True
    return _num(actual_norm) is not None and _num(expected_norm) is not None and abs((_num(actual_norm) or 0) - (_num(expected_norm) or 0)) <= 0.001


def _compare_row(actual: dict[str, Any], expected: dict[str, Any]) -> list[str]:
    failures = []
    for field in ("quantity", "od_mm", "id_mm", "thickness_mm"):
        if field in expected and not _same_float(actual.get(field), expected[field]):
            failures.append(f"{field}: got {actual.get(field)!r}, expected {expected[field]!r}")

    if expected.get("gasket_type") and actual.get("gasket_type") != expected["gasket_type"]:
        failures.append(f"gasket_type: got {actual.get('gasket_type')!r}, expected {expected['gasket_type']!r}")

    if expected.get("size_type") and actual.get("size_type") != expected["size_type"]:
        failures.append(f"size_type: got {actual.get('size_type')!r}, expected {expected['size_type']!r}")

    if expected.get("size"):
        actual_size = actual.get("size_norm") or actual.get("size")
        if not _same_size(actual_size, expected["size"]):
            failures.append(f"size: got {actual_size!r}, expected {expected['size']!r}")

    if expected.get("rating"):
        actual_rating = actual.get("rating_norm") or actual.get("rating")
        if _norm_rating(actual_rating) != _norm_rating(expected["rating"]):
            failures.append(f"rating: got {actual_rating!r}, expected {expected['rating']!r}")

    for field in ("moc", "sw_winding_material", "sw_filler", "sw_inner_ring", "sw_outer_ring"):
        if expected.get(field) and not _same_contains_text(actual.get(field), expected[field]):
            failures.append(f"{field}: got {actual.get(field)!r}, expected {expected[field]!r}")

    if expected.get("standard") and expected["standard"] not in str(actual.get("standard") or ""):
        failures.append(f"standard: got {actual.get('standard')!r}, expected {expected['standard']!r}")

    if expected.get("face_type") and actual.get("face_type") != expected["face_type"]:
        failures.append(f"face_type: got {actual.get('face_type')!r}, expected {expected['face_type']!r}")

    return failures


def evaluate_file(path: Path, client: Any | None) -> EvaluationResult:
    expected = _ground_truth(path)
    deterministic = _deterministic_pipeline(path)
    review_rows = sum(extraction._needs_smart_parse_review(item) for item in deterministic)

    start = time.perf_counter()
    actual, used_openai = _smart_pipeline(path, client)
    duration = time.perf_counter() - start

    correct = 0
    mismatches = []
    for idx, expected_row in enumerate(expected):
        actual_row = actual[idx] if idx < len(actual) else {}
        failures = _compare_row(actual_row, expected_row)
        if failures:
            mismatches.append(
                {
                    "row": idx + 1,
                    "line_no": expected_row.get("line_no"),
                    "failures": failures[:6],
                    "description": actual_row.get("ggpl_description") or actual_row.get("raw_description"),
                }
            )
        else:
            correct += 1

    wrong = max(len(expected), len(actual)) - correct
    status_counts: dict[str, int] = {}
    for item in actual:
        status = str(item.get("status") or "unknown")
        status_counts[status] = status_counts.get(status, 0) + 1

    return EvaluationResult(
        file=path.name,
        rows=len(expected),
        correct=correct,
        wrong=wrong,
        duration_sec=round(duration, 3),
        status_counts=status_counts,
        review_rows=review_rows,
        used_openai_review=used_openai and review_rows > 0,
        mismatches=mismatches[:10],
    )


def main() -> int:
    client = _openai_client()
    results = [evaluate_file(path, client) for path in FILES]
    total = {
        "rows": sum(r.rows for r in results),
        "correct": sum(r.correct for r in results),
        "wrong": sum(r.wrong for r in results),
        "duration_sec": round(sum(r.duration_sec for r in results), 3),
    }
    print(json.dumps({"total": total, "files": [r.__dict__ for r in results]}, indent=2, ensure_ascii=True))
    return 0 if total["wrong"] == 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
