from __future__ import annotations

import json
import os
import re
import sys
import time
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
PACKAGES = ROOT / "packages"
for path in (ROOT, PACKAGES):
    if str(path) not in sys.path:
        sys.path.insert(0, str(path))

from core.formatter import format_description
from core.parser import parse_excel_file
from core.rules import apply_rules
from services import extraction

try:
    from dotenv import load_dotenv
except Exception:  # pragma: no cover
    load_dotenv = None


def _norm_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().upper())


def _norm_description(value: Any) -> str:
    text = str(value or "").upper()
    text = text.replace("DESCREPTION", "DESCRIPTION")
    text = text.replace("STAINLESS STEEL", "SS")
    text = text.replace("FLEXIBLE GRAPHITE", "FG")
    text = text.replace("GRAPHITE FILLER", "FG")
    text = text.replace("FILLER", "")
    text = text.replace("WINDINGS", "WINDING")
    text = text.replace("WINDING", "")
    text = text.replace("NOM THK", "GASKET THK")
    text = text.replace("THK", "THK")
    text = text.replace("CL", "")
    text = text.replace("CLASS", "")
    text = text.replace("ALLOY825", "ALLOY 825")
    text = text.replace("INCOLOY825", "ALLOY 825")
    text = text.replace("INCOLOY 825", "ALLOY 825")
    text = text.replace("AISI316", "316")
    text = text.replace("AISI 316", "316")
    text = text.replace("SS316", "316")
    text = text.replace("SS 316", "316")
    text = text.replace("SS316L", "316L")
    text = text.replace("SS 316L", "316L")
    text = text.replace("CS OUTER RING", "CS OR")
    text = text.replace("CS-OR", "CS OR")
    text = text.replace("INNER RING", "IR")
    text = text.replace("-IR", " IR")
    text = text.replace("OUTER RING", "OR")
    text = text.replace("SPIRAL WOUND", "SPW")
    return re.sub(r"[^A-Z0-9#./\"]+", "", text)


def _tokenize(value: Any) -> set[str]:
    text = str(value or "").upper()
    replacements = {
        "STAINLESS STEEL": "SS",
        "FLEXIBLE GRAPHITE": "FG",
        "EXFOLIATED EXPANDED GRAPHITE": "EXFOLIATED_EXPANDED_GRAPHITE",
        "ASME B16.20": "ASME_B16.20",
        "SPIRAL WOUND": "SPW",
        "OUTER RING": "OR",
        "INNER RING": "IR",
        "INCOLOY 825": "ALLOY_825",
        "ALLOY 825": "ALLOY_825",
        "SS 316L": "316L",
        "SS316L": "316L",
        "SS 316": "316",
        "SS316": "316",
        "AISI 316": "316",
        "CS-OR": "CS OR",
        "-IR": " IR",
    }
    for old, new in replacements.items():
        text = text.replace(old, new)
    return set(re.findall(r'ASME_B16\.20|ALLOY_825|EXFOLIATED_EXPANDED_GRAPHITE|316L|316|CS|FG|SPW|IR|OR|\d+(?:\.\d+)?"|\d+/?\d*#?', text))


def _norm_component(value: Any) -> str:
    text = str(value or "").upper()
    text = text.replace("SS316/316L", "316/316L")
    text = text.replace("SS316L", "316L").replace("SS 316L", "316L")
    text = text.replace("SS316", "316").replace("SS 316", "316")
    text = text.replace("SS304L", "304L").replace("SS 304L", "304L")
    text = text.replace("SS304", "304").replace("SS 304", "304")
    text = text.replace("FLEXIBLE GRAPHITE", "FG").replace("GRAPHITE", "FG")
    text = text.replace("INCOLOY 825", "ALLOY 825").replace("INCOLOY825", "ALLOY 825")
    text = text.replace("EXFOLIATED EXPANDED FG", "EXFOLIATED EXPANDED GRAPHITE")
    text = text.replace("EXPANDED FG", "EXPANDED GRAPHITE")
    return re.sub(r"[^A-Z0-9/#.\"]+", "", text)


def _parse_expected_spw(value: Any) -> dict[str, str]:
    text = str(value or "").upper()
    if "SPW" not in text:
        return {}
    out: dict[str, str] = {}
    head = re.search(r"SPW\s+(.+?)X\s*([0-9/]+#)", text)
    if head:
        out["size"] = head.group(1).strip()
        out["rating"] = head.group(2).strip()
    outer = re.search(r"([A-Z0-9 ./-]+?)-OR\+", text)
    if outer:
        out["sw_outer_ring"] = outer.group(1).strip()
    body = re.search(r"-OR\+(.+?)\s+\d+(?:\.\d+)?\s*MM\s+GASKET", text)
    if body:
        construction = body.group(1).strip()
        inner = re.search(r"\+(.+?)-IR\b", construction)
        if inner:
            out["sw_inner_ring"] = inner.group(1).strip()
            construction = construction[: inner.start()]
        if "/" in construction:
            winding, filler = construction.split("/", 1)
            out["sw_winding_material"] = winding.strip()
            out["sw_filler"] = filler.strip()
    return out


def _component_score(actual_items: list[dict[str, Any]], expected_rows: list[dict[str, Any]]) -> dict[str, Any]:
    fields = ["size", "rating", "sw_winding_material", "sw_filler", "sw_inner_ring", "sw_outer_ring"]
    counts = {field: {"correct": 0, "total": 0} for field in fields}
    full = 0
    rows = 0
    ir_or_full = 0
    ir_or_rows = 0

    for idx, expected_row in enumerate(expected_rows):
        expected = _parse_expected_spw(expected_row.get("expected"))
        if not expected:
            continue
        rows += 1
        actual = actual_items[idx] if idx < len(actual_items) else {}
        actual_values = {
            "size": actual.get("size_norm") or actual.get("size"),
            "rating": actual.get("rating"),
            "sw_winding_material": actual.get("sw_winding_material"),
            "sw_filler": actual.get("sw_filler"),
            "sw_inner_ring": actual.get("sw_inner_ring"),
            "sw_outer_ring": actual.get("sw_outer_ring"),
        }
        row_ok = True
        for field in fields:
            if field not in expected:
                continue
            counts[field]["total"] += 1
            ok = _norm_component(actual_values.get(field)) == _norm_component(expected[field])
            if ok:
                counts[field]["correct"] += 1
            else:
                row_ok = False
        if "sw_inner_ring" in expected and "sw_outer_ring" in expected:
            ir_or_rows += 1
            if (
                _norm_component(actual_values.get("sw_inner_ring")) == _norm_component(expected["sw_inner_ring"])
                and _norm_component(actual_values.get("sw_outer_ring")) == _norm_component(expected["sw_outer_ring"])
            ):
                ir_or_full += 1
        if row_ok:
            full += 1

    by_field = {
        field: {
            **counts[field],
            "accuracy_pct": round(counts[field]["correct"] / counts[field]["total"] * 100, 2)
            if counts[field]["total"]
            else 0,
        }
        for field in fields
    }
    return {
        "spw_rows_scored": rows,
        "full_component_rows": full,
        "full_component_accuracy_pct": round(full / rows * 100, 2) if rows else 0,
        "ir_or_rows": ir_or_rows,
        "ir_or_both_correct": ir_or_full,
        "ir_or_accuracy_pct": round(ir_or_full / ir_or_rows * 100, 2) if ir_or_rows else 0,
        "by_field": by_field,
    }


def _load_rows(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Master"] if "Master" in wb.sheetnames else wb.active
    headers = [_norm_header(cell.value) for cell in ws[1]]
    col = {header: idx for idx, header in enumerate(headers) if header}

    def get(row: tuple, name: str) -> Any:
        idx = col.get(_norm_header(name))
        return row[idx] if idx is not None and idx < len(row) else None

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        expected = get(row, "GGPL DESCREPTION")
        desc = get(row, "MATERIAL DESCRIPTION")
        sl_no = get(row, "SL.NO.")
        if not expected and not desc:
            continue
        rows.append(
            {
                "sl_no": sl_no,
                "line_number": get(row, "LINE NUMBER"),
                "material": get(row, "MATERIAL"),
                "material_description": desc,
                "size": get(row, "SIZE (INCH)"),
                "uom": get(row, "UOM"),
                "qty": get(row, "QTY"),
                "unit_price": get(row, "UNIT PRICE"),
                "expected": expected,
            }
        )
    return rows


def _processed_fast(path: Path) -> list[dict[str, Any]]:
    processed = []
    for raw in parse_excel_file(path.read_bytes()):
        item = apply_rules(dict(raw))
        item["ggpl_description"] = format_description(item)
        processed.append(item)
    return processed


def _client() -> Any | None:
    if os.getenv("SMART_EVAL_NO_OPENAI"):
        return None
    if load_dotenv:
        load_dotenv(ROOT / ".env")
    if not os.getenv("OPENAI_API_KEY"):
        return None
    from openai import OpenAI

    return OpenAI(timeout=180.0)


def _score(actual_items: list[dict[str, Any]], expected_rows: list[dict[str, Any]]) -> dict[str, Any]:
    exact = 0
    normalized_exact = 0
    token_recalls: list[float] = []
    token_precisions: list[float] = []
    full_expected_token_coverage = 0
    examples = []

    for idx, expected_row in enumerate(expected_rows):
        actual = actual_items[idx] if idx < len(actual_items) else {}
        actual_desc = actual.get("ggpl_description") or actual.get("raw_description") or ""
        expected_desc = expected_row.get("expected") or ""

        if str(actual_desc).strip() == str(expected_desc).strip():
            exact += 1
        if _norm_description(actual_desc) == _norm_description(expected_desc):
            normalized_exact += 1

        actual_tokens = _tokenize(actual_desc)
        expected_tokens = _tokenize(expected_desc)
        if expected_tokens:
            recall = len(actual_tokens & expected_tokens) / len(expected_tokens)
            token_recalls.append(recall)
            if recall == 1:
                full_expected_token_coverage += 1
        if actual_tokens:
            token_precisions.append(len(actual_tokens & expected_tokens) / len(actual_tokens))

        if len(examples) < 12 and _norm_description(actual_desc) != _norm_description(expected_desc):
            examples.append(
                {
                    "row": idx + 2,
                    "sl_no": expected_row.get("sl_no"),
                    "status": actual.get("status"),
                    "flags": (actual.get("flags") or [])[:3],
                    "expected": expected_desc,
                    "actual": actual_desc,
                }
            )

    row_count = len(expected_rows)
    return {
        "row_count": row_count,
        "pipeline_rows": len(actual_items),
        "exact_description_matches": exact,
        "exact_description_accuracy_pct": round(exact / row_count * 100, 2) if row_count else 0,
        "normalized_description_matches": normalized_exact,
        "normalized_description_accuracy_pct": round(normalized_exact / row_count * 100, 2) if row_count else 0,
        "avg_expected_token_recall_pct": round(sum(token_recalls) / len(token_recalls) * 100, 2) if token_recalls else 0,
        "avg_actual_token_precision_pct": round(sum(token_precisions) / len(token_precisions) * 100, 2) if token_precisions else 0,
        "rows_with_all_expected_tokens": full_expected_token_coverage,
        "all_expected_tokens_accuracy_pct": round(full_expected_token_coverage / row_count * 100, 2) if row_count else 0,
        "sample_mismatches": examples,
    }


def main() -> int:
    if len(sys.argv) != 2:
        print("Usage: python scripts/analyze_axiom_excel.py <xlsx-path>", file=sys.stderr)
        return 2

    path = Path(sys.argv[1])
    expected_rows = _load_rows(path)

    fast_start = time.perf_counter()
    fast_items = _processed_fast(path)
    fast_sec = time.perf_counter() - fast_start
    review_rows = sum(extraction._needs_smart_parse_review(item) for item in fast_items)

    client = _client()
    prod_start = time.perf_counter()
    if client:
        items, skipped, error = extraction.process_document(path.read_bytes(), "excel", client)
        openai_client_present = True
    else:
        items, skipped, error = fast_items, 0, "OPENAI_API_KEY not available; used deterministic fast path only"
        openai_client_present = False
    prod_sec = time.perf_counter() - prod_start

    changed_by_review = sum(
        1
        for before, after in zip(fast_items, items)
        if (before.get("ggpl_description"), before.get("status"), before.get("flags"))
        != (after.get("ggpl_description"), after.get("status"), after.get("flags"))
    )

    output = {
        "file": str(path),
        "input_rows_with_expected": len(expected_rows),
        "fast_path_rows": len(fast_items),
        "review_rows_flagged_before_openai": review_rows,
        "openai_client_present": openai_client_present,
        "production_error": error,
        "skipped": skipped,
        "timings_sec": {
            "fast_path_parse_rules_format": round(fast_sec, 3),
            "production_process_document_total": round(prod_sec, 3),
        },
        "status_counts": dict(Counter(item.get("status") for item in items)),
        "changed_by_smart_review_rows": changed_by_review,
        "fast_path_accuracy": _score(fast_items, expected_rows),
        "fast_path_component_accuracy": _component_score(fast_items, expected_rows),
        "accuracy": _score(items, expected_rows),
        "component_accuracy": _component_score(items, expected_rows),
    }
    print(json.dumps(output, indent=2, ensure_ascii=True))
    return 0 if not error and items else 1


if __name__ == "__main__":
    raise SystemExit(main())
