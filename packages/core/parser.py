from __future__ import annotations
"""
Parses customer enquiry inputs (email text or Excel) into a uniform list of raw items:
[{'line_no': int, 'description': str, 'quantity': float, 'uom': str}]
"""
import re
import io
import logging
import openpyxl

logger = logging.getLogger(__name__)


def worksheet_rows_with_merged_values(ws, max_row: int | None = None) -> list[tuple]:
    """Return worksheet rows with merged cells expanded to every covered row.

    openpyxl exposes only the top-left value of a merged range. Enquiry sheets
    often merge MOC/spec cells vertically, so reading values directly makes the
    following line items look incomplete.
    """
    row_limit = min(max_row or ws.max_row, ws.max_row)
    rows = [
        [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, row_limit + 1)
    ]

    for merged_range in ws.merged_cells.ranges:
        value = ws.cell(merged_range.min_row, merged_range.min_col).value
        if value is None:
            continue
        for r in range(merged_range.min_row, min(merged_range.max_row, row_limit) + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                rows[r - 1][c - 1] = value

    return [tuple(row) for row in rows]


def parse_email_text(text: str) -> list[dict]:
    """Extract line items from pasted email body text via regex/rule-based parsing."""
    items = []
    lines = _merge_continuation_lines([l.strip() for l in text.splitlines() if l.strip()])
    for line in lines:
        item = _parse_line(line)
        if item:
            items.append(item)
    # Number items if line_no missing
    for i, item in enumerate(items, 1):
        if not item.get('line_no'):
            item['line_no'] = i
    return items


def _merge_continuation_lines(lines: list[str]) -> list[str]:
    """Merge lines that are mid-sentence continuations of the previous line.

    Handles cases where a long gasket description wraps across two lines,
    e.g. Excel cell content copied as text where a description ends with
    'OUTER RING' and the next line starts with 'MATERIAL: Carbon Steel...'.
    """
    merged = []
    for line in lines:
        if not merged:
            merged.append(line)
            continue
        prev = merged[-1]
        # A line is a continuation if:
        # 1. It does NOT start with a digit (new serial number), AND
        # 2. The previous line ended with a comma, OR ended with RING/MATERIAL/OUTER/INNER
        #    (cut mid-phrase), OR the current line starts with a field-label ("WORD:") pattern
        starts_with_number = bool(re.match(r'^\d', line))
        # Lines starting with a gasket size-prefix keyword are always new items
        is_new_item_prefix = bool(re.match(r'^(?:NPS|NB|DN|SIZE)\s*:', line, re.IGNORECASE))
        prev_ends_mid = (
            prev.endswith(',')
            or re.search(r'\b(?:OUTER\s*RING|INNER\s*RING|OUTER|RING|MATERIAL)\s*$', prev, re.IGNORECASE)
        )
        # Previous line ends with a standards-body acronym — next line is its number (e.g. "ANSI\n B16.47")
        prev_ends_std_prefix = bool(
            re.search(r'\b(?:ANSI|ASME|API|ISO|EN|DIN|BS|ASTM|NACE|IBR|AWS)\s*$', prev, re.IGNORECASE)
        )
        curr_is_field_continuation = bool(re.match(r'^[A-Z][A-Z\s]+:\s*\S', line))
        # Current line is a standard reference line (e.g. "ASME B16.20: Metallic Gaskets for Pipe Flanges")
        # that belongs to the preceding description, not a new line item
        curr_is_standard_ref = bool(re.match(
            r'^(?:ASME|ANSI|API|ISO|EN|DIN|BS|ASTM|NACE|IBR|AWS)\b', line, re.IGNORECASE
        ))
        if (not starts_with_number and not is_new_item_prefix
                and (prev_ends_mid or prev_ends_std_prefix or curr_is_field_continuation or curr_is_standard_ref)):
            merged[-1] = prev + ' ' + line
        else:
            merged.append(line)
    return merged


def _parse_line(line: str) -> "dict | None":
    """Try to extract (description, qty, uom) from a single text line."""
    # Skip header-like lines
    lower = line.lower()
    if any(kw in lower for kw in ['sl.no', 'sl no', 'line no', 'release no', 'notes', 'inv uom',
                                   'description', 'subject', 'dear', 'regards', 'kindly',
                                   'total', 'terms', 'price', 'revision', 'date']):
        return None

    # Pattern: optional number, description text, number, unit
    # e.g. "1  Gasket - Rubber - 6'' PN10  27  m"
    pattern = re.compile(
        r'^(\d+)?\s*'                            # optional sl.no
        r'(?:\d+\s+\d+\s+)?'                     # optional line/release no columns
        r'(.+?)\s+'                              # description (greedy)
        r'(\d+(?:\.\d+)?)\s*'                    # quantity
        r'(nos?|m|mtr|meters?|pcs?|sets?|kgs?|units?)\s*$',  # unit
        re.IGNORECASE
    )
    m = pattern.match(line)
    if m:
        sl_no = int(m.group(1)) if m.group(1) else None
        desc = m.group(2).strip()
        qty = float(m.group(3))
        uom = _normalize_uom(m.group(4))
        if _looks_like_gasket(desc) and qty > 0:
            return {'line_no': sl_no, 'description': desc, 'quantity': qty, 'uom': uom}

    # Simpler fallback: tab/multiple-space separated columns
    parts = re.split(r'\t|  +', line)
    parts = [p.strip() for p in parts if p.strip()]
    if len(parts) >= 3:
        # Try last two parts as qty + uom, rest as description
        try:
            qty = float(parts[-2])
            uom = _normalize_uom(parts[-1])
            desc = ' '.join(parts[:-2]).strip()
            desc = re.sub(r'^\d+\s+', '', desc).strip()
            if _looks_like_gasket(desc) and qty > 0:
                return {'line_no': None, 'description': desc, 'quantity': qty, 'uom': uom}
        except (ValueError, IndexError):
            pass

    # Last resort: entire line is a description with no quantity
    # Strip leading sl.no if present: "1. " / "1) " — but NOT "1.5" or "1.1/2" (period followed by digit)
    desc = re.sub(r'^\d+[\.\)](?!\d)\s*', '', line).strip()
    if _looks_like_gasket(desc):
        return {'line_no': None, 'description': desc, 'quantity': None, 'uom': 'NOS'}

    return None


def _looks_like_gasket(text: str) -> bool:
    text_lower = text.lower()
    # Must contain gasket-related keyword or size indicator
    gasket_kws = ['gasket', 'gkt', 'rubber', 'ptfe', 'neoprene', 'epdm', 'cnaf',
                  'viton', 'graphite', 'grph', 'graph fill', 'pn', '150#', '300#', '600#',
                  'asme', 'ansi', 'b16.20', 'b16.21',
                  'rtj', 'r.t.j', 'ring joint', 'joint tore', 'tore', 'spiral', 'winding',
                  'spw', 'wnd', 'sw gasket', 'kammprofile', 'kamprofile', 'camprofile', 'insulating gasket',
                  'insulation gasket', 'isolating kit', 'insulating kit',
                  'isk', 'soft iron', 'softiron', 'octagonal', 'oval ring',
                  'nbr', 'nitrile', 'sbr', 'silicone', 'butyl', 'aramid', 'thermiculite',

                  'expanded graphite', 'cork', 'leather', 'ceramic fiber', 'hnbr',
                  'outer ring', 'inner ring', 'centering ring', 'gid', 'god']
    has_kw = any(k in text_lower for k in gasket_kws)
    has_size = bool(re.search(
        r'\d+["\']|\d+\s*(?:nb|dn|nps|inch|mm)|(?:nb|dn)\s*\d+|\d+\s*(?:gid|god)\b',
        text_lower, re.IGNORECASE))
    return has_kw or has_size


def _normalize_uom(uom: str) -> str:
    uom = uom.strip().upper()
    if uom.startswith('M') and not uom.startswith('MT'):
        return 'M'
    return 'NOS'


def parse_excel_file(file_bytes: bytes) -> list[dict]:
    """Parse an uploaded Excel file into raw line items via rule-based detection."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_items = _parse_sheet(ws)
        if sheet_items:
            items.extend(sheet_items)
    for i, item in enumerate(items, 1):
        if not item.get('line_no'):
            item['line_no'] = i
        item.setdefault('source_index', i)
    return items


def excel_requires_smart_parse(file_bytes: bytes) -> bool:
    """Return True for layouts where the app should not use Excel fast path."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if any(rng.min_row != rng.max_row for rng in ws.merged_cells.ranges):
            return True

        rows = worksheet_rows_with_merged_values(ws)
        description_sections = _detect_description_sections(rows)
        structured_sections = _detect_structured_sections(rows)
        if len(description_sections) + len(structured_sections) > 1:
            return True

    return False


def _parse_sheet(ws) -> list[dict]:
    """Detect header row and extract line items from a worksheet (rule-based)."""
    items = _parse_description_sections(ws)
    if items:
        return items
    return _parse_structured_sheet(ws)


def _norm_header_cell(cell) -> str:
    """Normalize a header cell: lowercase, collapse whitespace incl. non-breaking spaces."""
    if cell is None:
        return ''
    return re.sub(r'[\xa0\s]+', ' ', str(cell)).strip().lower()


def _classify_structured_col(norm: str) -> str | None:
    """Return the column type for a normalized header cell, or None if unrecognised."""
    if norm in ('dn', 'nb'):
        return 'dn_size'
    if norm == 'class':
        return 'class_rating'
    if norm in ('material', 'moc'):
        return 'material'
    if norm in ('thickness', 'thk', 'thk (mm)', 'thick'):
        return 'thickness'
    if norm.startswith('(od)') or norm.startswith('od (') or norm.startswith('od mm') or norm == 'od':
        return 'od_mm'
    if norm.startswith('(id)') or norm.startswith('id (') or norm.startswith('id mm') or norm == 'id':
        return 'id_mm'
    # SIZE column (inches): "SIZE", "SIZE (INCH)", "SIZE\n(INCH)", "SIZE (MM)", "NPS"
    if norm in ('size', 'nps') or norm.startswith('size (') or norm.startswith('size\n'):
        return 'size_inch'
    # RATING column: "RATING", "RATING ", "PRESSURE RATING", "CLASS/RATING"
    if norm in ('rating', 'pressure rating', 'class/rating', 'rating (class)') or norm.startswith('rating'):
        return 'rating'
    if norm in ('qty', 'quantity'):
        return 'quantity'
    if norm in ('uom', 'inv uom'):
        return 'uom'
    if 'sl.no' in norm or 'sr.no' in norm or norm == 'sno' or norm == 'sl no' or norm == 'sr no':
        return 'line_no'
    return None


def _classify_description_col(norm: str) -> str | None:
    if not norm:
        return None
    if norm in ('piping class', 'pipe class'):
        return None
    if re.search(r'\bpr\b', norm) and re.search(r'\bsr\.?\s*no\.?\b', norm):
        return None
    for col_type, keywords in _HEADER_PATTERNS.items():
        if any(kw in norm for kw in keywords):
            return col_type
    if norm == 'material':
        return 'moc'
    return _classify_structured_col(norm)


def _detect_description_sections(all_rows: list[tuple]) -> list[tuple]:
    """Find all description-column header blocks in a sheet."""
    sections = []
    current_header_idx = None
    current_col_map = None
    current_data: list[tuple] = []

    def _is_description_header(col_map: dict) -> bool:
        has_description_source = (
            'description' in col_map
            or 'technical_description' in col_map
            or 'dimension' in col_map
            or 'gasket_form' in col_map
        )
        return has_description_source and (
            'quantity' in col_map or 'moc' in col_map or 'size_inch' in col_map
            or 'rating' in col_map or ('od_mm' in col_map and 'id_mm' in col_map)
        )

    for row_idx, row in enumerate(all_rows):
        col_map = {}
        col_priority = {}
        for col_idx, cell in enumerate(row):
            norm = _norm_header_cell(cell)
            col_type = _classify_description_col(norm)
            if not col_type:
                continue
            priority = _header_priority(col_type, norm)
            if col_type not in col_map or priority > col_priority.get(col_type, 0):
                col_map[col_type] = col_idx
                col_priority[col_type] = priority

        if _is_description_header(col_map):
            if current_col_map is not None and current_data:
                sections.append((current_header_idx, current_col_map, current_data))
            current_header_idx = row_idx
            current_col_map = col_map
            current_data = []
        elif current_col_map is not None and any(c is not None for c in row):
            current_data.append(row)

    if current_col_map is not None and current_data:
        sections.append((current_header_idx, current_col_map, current_data))

    return sections


def _header_priority(col_type: str, norm: str) -> int:
    if col_type == 'quantity':
        if any(token in norm for token in ('to be purchased', 'balance to order', 'required qty', 'qty with margin')):
            return 30
        if 'req qty' in norm or 'required' in norm:
            return 20
        return 10
    return 10


def _append_field(parts: list[str], label: str, value: str | None, suffix: str = '') -> None:
    if value:
        if re.fullmatch(r'n/?a|not applicable|nil|none', value.strip(), re.IGNORECASE):
            return
        parts.append(f'{label}: {value}{suffix}')


def _float_from_text(value: str | None) -> float | None:
    if not value:
        return None
    match = re.search(r'\d+(?:\.\d+)?', str(value))
    return float(match.group(0)) if match else None


def _rating_from_text(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip().upper()
    raw = re.sub(r'(PN)\s*(\d+)', r'\1 \2', raw)
    # PN classes run 2.5–100 for flanges and up to PN400 for lens rings.
    # "PN:{1-9}" adjacent to drawing/part context is a PART number (Rule O).
    pn = re.search(r'\bPN\s*-?\s*(\d{1,3})(?!\d)', raw)
    if pn and not (len(pn.group(1)) == 1 and re.search(r'DRAWING|PART|\bDIA\b', raw)):
        return f'PN {pn.group(1)}'
    lb = re.search(r'\b(150|300|400|600|900|1500|2500)\s*LB\b', raw)
    if lb:
        return f'{lb.group(1)}#'
    dual_cls = re.search(
        r'\b(?:CL(?:ASS)?\.?|#)?\s*(150|300|400|600|900|1500|2500)\s*#?\s*'
        r'(?:/|\\|OR|AND|-)\s*'
        r'(?:CL(?:ASS)?\.?|#)?\s*(150|300|400|600|900|1500|2500)\s*#?\b',
        raw,
    )
    if dual_cls:
        return f'{dual_cls.group(1)}/{dual_cls.group(2)}#'
    cls = re.search(r'\b(?:CL(?:ASS)?\.?|#)?\s*(150|300|400|600|900|1500|2500)\s*#?\b', raw)
    if cls:
        return f'{cls.group(1)}#'
    return raw if len(raw) <= 20 else None


def _size_from_text(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).strip()
    fraction_map = {
        '¼': '0.25',
        '½': '0.5',
        '¾': '0.75',
    }
    if raw in fraction_map:
        raw = fraction_map[raw]
    try:
        number = float(raw)
        return f'{int(number)}"' if number == int(number) else f'{number}"'
    except ValueError:
        return raw


# RULE J-2 Part 1 — spaced/period-broken abbreviations ("S P W D", "R.T.J")
# left behind by manual typing, OCR, and PDF text extraction. A run of 2–5
# single letters is collapsed ONLY when the joined form is a known gasket
# abbreviation, so real words ("SS INNER") and dimensions ("4 MM") are never
# touched.
_SPACED_ABBREV_CANONICAL = {
    'SPWD', 'SPW', 'SWG', 'RTJ', 'CNAF', 'IR', 'OR', 'CR', 'DJ', 'DJI',
    'MCR', 'FF', 'RF',
}
_SPACED_ABBREV_RE = re.compile(
    r'(?<![A-Za-z0-9.])([A-Za-z](?:[ .][A-Za-z]){1,4})(?![A-Za-z0-9])'
)


def _collapse_spaced_abbrevs(text: str) -> str:
    def _sub(m: re.Match) -> str:
        joined = re.sub(r'[ .]', '', m.group(1)).upper()
        return joined if joined in _SPACED_ABBREV_CANONICAL else m.group(0)
    return _SPACED_ABBREV_RE.sub(_sub, text)


def _infer_gasket_type(description: str) -> str | None:
    raw = _collapse_spaced_abbrevs(description).upper()
    raw = re.sub(r'(?i)(GASKET)(?=(SKAG|CAM|KAMM|DOUBLE|COPPER|\d))', r'\1 ', raw)
    raw = re.sub(r'(?i)(\d)(INST\.?\s+KIT|INSULATING|IN\s+GASKET)', r'\1 \2', raw)
    # Adjacent (non-gasket) products — quoted as REGRET per GGPL policy
    if re.search(
        r'SPECTACLE\s+BLIND|SPADE\s*(?:&|AND)\s*SPACER|PADDLE\s+BLANK|'
        r'EXPANSION\s+JOINT|BELLOW(?:S)?\b|THERMAL\s+INSULATION|INSULATION\s+(?:MATERIAL|CLOTH|ROPE)|'
        r'GLAND\s+PACKING|BRAIDED\s+PACKING', raw,
    ):
        return 'ADJACENT'
    if re.search(r'\bSTUD(?:\s+BOLT)?S?\b', raw) and re.search(r'\bNUTS?\b|ASTM\s+A19[34]', raw) \
            and not re.search(r'\bGASKET\b|\bGSKT\b|\bISK\b|INSULAT', raw):
        return 'ADJACENT'
    if re.search(r'\bLENS\s+(?:RING|GASKET)\b|\bLENTICULAR\b', raw):
        return 'LENS'
    if re.search(r'\bMAN\s?HOLE\b|\bHAND\s?HOLE\b', raw) and re.search(r'\bGASKET\b|\bGSKT\b', raw):
        return 'MANHOLE'
    if re.search(r'\bENVELOPE\s+GASKET\b|PTFE\s+ENVELOPE', raw):
        return 'ENVELOPE'
    if re.search(
        r'INSULAT(?:ING|ION)|\bNSULATING\b|\bISK\b|\bFLANGE\s+ISOLATION\b|'
        r'\bFLANGE\s+INSULATION\s+KIT\b|\bFLANGE\s+ISOLAT(?:ING|ION)\s+KIT\b|'
        r'\bINSULATION\s+KIT\b|\bISOLAT(?:ING|ION)\s+GASKET\b|\bCOMMANDER\s+EXTREME\b|\bINST\.?\s+KIT\b',
        raw,
    ):
        if re.search(r'\bRTJ\b|\bR/?J\b|\bRING\s+JOINT\b|\bTYPE[-\s]?D\b', raw):
            return 'ISK_RTJ'
        return 'ISK'
    if re.search(
        r'\b(?:SPIRAL|SPRIAL|SPRIRAL|SPIRIAL|SPLRAL|SPRLAL|SPIRRAL|SPRRAL|SPRL|SPIR|SPL|SP)\s*[-\s]*(?:W(?:OU)?ND\w*|WIND\w*)\b'
        r'|\bSPW[DG]?\b|\bSW\s+GASKET\b|\bSWG\b|\bGASKET\s*,?\s*SW\b|\bGASW\b'
        r'|\bGASKET\s+SPIRAL\b|\bSPIRAL\s+GASKET\b|\bJOINT\s+SPIRALE?\b'
        r'|\bRESTRICTION\s+ORIFICE\b',
        raw,
    ):
        return 'SPIRAL_WOUND'
    # "MOC :- SS316 / GRAPHITE FILLER + SS316 IR & CS OR" style rows are SPW:
    # a filler plus inner/outer ring components only exist on spiral wound.
    # (kammprofile/jacketed/RTJ cues take precedence — checked in the guard.)
    if (re.search(r'\bGASKET\b', raw) and re.search(r'FILLER|FILLED', raw)
            and re.search(r'\bI\.?R\.?\b|\bO\.?R\.?\b|\bINNER\b|\bOUTER\b|CENTERING|CENTRING', raw)
            and not re.search(r'KAMM|KAMPROFILE|CAMPROFILE|\bKMP\b|PROFILE|JACKET|SKAG|\bRTJ\b|RING\s+JOINT|GROOVED', raw)):
        return 'SPIRAL_WOUND'
    if re.search(r'\b(?:RING\s+JOINT|RING\s+TYPE\s+JOINT|RING\s+TYPE\s+GASKET|RTJ|RJ\s+GASKET|R/?J)\b|'
                 r'\b(?:OCTAGONAL|OVAL)\s+RING\b|'
                 r'\b(?:BX|RX)\s*-?\s*1?\d{2}\b', raw):
        return 'RTJ'
    if re.search(r'\bRING\s+(?:GASKET|GSKT)S?\b|\bGSKT\s*,?\s*R(?:I|L)?NG\b|\bRNG\s*,?\s*GSKT\b|\bRING\s+GSKT\b', raw) and re.search(
            r'\bOCTAGONAL\b|\bOVAL\b|\bSOFT\s+IRON\b|\bAPI\b|\bBHN\b|\bBOP\b|'
            r'\bRING\s+(?:IDENTIFICATION\s+)?(?:NUMBER|NO)\b|\bR\s*-\s*\d{1,3}\b', raw):
        return 'RTJ'
    if re.search(r'\bR\s*-\s*\d{1,3}\b', raw) and re.search(r'\bGASKET\b|\bGSKT\b|\bRNG\b', raw):
        return 'RTJ'
    # RULE Y Part 1 — KAMM detection (ISK and RTJ excepted, KAMM wins over SPW).
    # `PROFILE GASKET` is ALWAYS KAMM, even with an FKM/rubber facing.
    # `MET GRVD` / `GRVD` in a short text = grooved metal = KAMM.
    if re.search(
        r'\bKAMMPROFILE\b|\bKAMPROFILE\b|\bKAMM\s*PROFILE\b|\bCAMPROFILE\b|\bCAM\s*PROFILE\b'
        r'|\bPROFILE\s+GASKET\b|\bGROOVED\s+PROFILE\b|\bGROOVED\s+METAL\b|SKAG|\bKMP\b'
        r'|\bSERRATED\s+METAL\b|\bSERRATED\b|\bKAMM\b|\bCAMP\b'
        r'|\bMET\s*GRVD\b|\bGRVD\b|\bMETAL\s+GROOVED\b'
        # Brand styles that denote a grooved-metal construction
        r'|\bFLEXPRO\b|\bKAMMPRO\b|\bMAXIPROFILE\b|\bLEADER[-\s]?KAM\b|\bMETAKAMM\b',
        raw,
    ):
        return 'KAMM'
    if re.search(r'\bDOUBLE[\s\-]?JACKET(?:ED)?\b|\bJACKETED\b|\bJACKET\s+GASKET\b|\bCOPPER\s+JACKET\b', raw):
        return 'DJI'
    if re.search(r'\bPLUG\s+GASKET\b|\bPLUG\s+TYPE\s+GASKET\b', raw):
        return 'PLUG_GASKET'
    # Standalone corrugated gasket = corrugated metal gasket (CMG family);
    # "corrugated type ... filler" inside a DJ text stays DJI (checked above)
    if re.search(r'\bCORRUGATED(?:\s+METAL(?:LIC)?)?\s+GASKET\b|\bCORRUGATED\s+GASKET\b|\bCMG\b'
                 r'|\b\d-?PLY\s+CORRUGATED\b|METAL\s+SOLID\s+GASKET\s+CORRUGATED', raw):
        return 'CMG'
    if re.search(r'\bMETAL\s+CLAD(?:DED)?\b|\bCLAD\s+GASKET\b', raw) and 'DOUBLE' not in raw:
        return 'METAL_CLAD'
    if re.search(r'\bSOLID\s+METAL\b|\bMETAL\s+FLAT\s+RING\b', raw):
        return 'SOLID_METAL'
    if re.search(r'\bLIP\s+SEAL\b', raw):
        return 'LIP_SEAL'
    if re.search(r'\bDIAPHRAG?M\b|\bDIAPHRAM\b', raw):
        return 'DIAPHRAGM'
    if re.search(r'\bEYELET', raw):
        return 'EYELET'
    if re.search(r'\bSHEET\s+GASKET\b|\bGASKET\s+SHEET\b', raw):
        return 'SHEET_GASKET'
    # Raw sheet/roll supply: SHEET/ROLL wording + L×W dims, no flange class
    if re.search(r'\bSHEETS?\b|\bROLLS?\b|JOINTING\s+SHEET', raw) \
            and re.search(r'\d\s*(?:MTR|M|MM)?\s*[X×]\s*\d', raw) \
            and not re.search(r'\b(?:150|300|400|600|900|1500|2500)\s*#|\bCL(?:ASS)?\.?\s*\d|\bPN\s*\d', raw):
        return 'SHEET'
    if re.search(r'\bGASKET\b.*\bO\.?D\.?\s*\d+.*\bI\.?D\.?\s*\d+', raw, re.IGNORECASE) and re.search(r'\bDRAWING\b|\bPOSITION\b|\bASBESTOS\s+FREE\b', raw):
        return 'DJI'
    if _looks_like_gasket(description):
        return 'SOFT_CUT'
    return None


def _standard_from_text(value: str | None) -> str | None:
    if not value:
        return None
    raw = str(value).upper()
    # RULE V legacy-typo tolerance: "ASME B16..20" (double dot), "B-16.21",
    # "ASME 16.20" (missing B) all normalize to the canonical ASME form.
    match = re.search(r'\b(?:(?:ASME|ANSI)\s*B?|B)[\s.-]*16\s*\.{1,2}\s*(20|21|47)\b', raw)
    if match:
        std = f'ASME B16.{match.group(1)}'
        if match.group(1) == '47':
            series = re.search(r'SERIES\s*[-:]?\s*([AB])\b', raw)
            if series:
                std = f'ASME B16.47 (SERIES-{series.group(1)})'
        return std
    match = re.search(r'\bAPI\s*6A\b', raw)
    if match:
        return 'API 6A'
    # Obsolete standards → successors (deviation note added by the rules engine)
    if re.search(r'\bAPI\s*601\b', raw):
        return 'ASME B16.20'
    if re.search(r'\bAPI\s*605\b', raw):
        return 'ASME B16.47 (SERIES-B)'
    if re.search(r'\bMSS\s*SP[-\s]?44\b', raw):
        return 'ASME B16.47 (SERIES-A)'
    match = re.search(r'\bEN\s*1514[-\s]*(\d+)\b', raw)
    if match:
        return f'EN 1514-{match.group(1)}'
    return None


_SW_MATERIAL_ALIASES: list[tuple[str, str]] = [
    (r'SS\s*TP\s*316\s*/\s*316L', 'SS316/SS316L'),
    (r'TP\s*316\s*/\s*316L', 'SS316/SS316L'),
    (r'SS\s*316\s*/\s*SS\s*316L', 'SS316/SS316L'),
    (r'SS\s*316\s*/\s*316L', 'SS316/SS316L'),
    (r'S\.?\s*S\.?\s*316L', 'SS316L'),
    (r'S\.?\s*S\.?\s*316', 'SS316'),
    (r'S\.?\s*S\.?\s*304L', 'SS304L'),
    (r'S\.?\s*S\.?\s*304', 'SS304'),
    (r'TP\s*316L\s*SS', 'SS316L'),
    (r'TP\s*316\s*SS', 'SS316'),
    (r'TP\s*304L\s*SS', 'SS304L'),
    (r'TP\s*304\s*SS', 'SS304'),
    (r'316L\s*SS', 'SS316L'),
    (r'316\s*SS', 'SS316'),
    (r'304L\s*SS', 'SS304L'),
    (r'304\s*SS', 'SS304'),
    (r'STAINLESS\s+STEEL\s+316L', 'SS316L'),
    (r'STAINLESS\s+STEEL\s+316', 'SS316'),
    (r'STAINLESS\s+STEEL\s+304L', 'SS304L'),
    (r'STAINLESS\s+STEEL\s+304', 'SS304'),
    (r'316L\s+STAINLESS\s+STEEL', 'SS316L'),
    (r'316\s+STAINLESS\s+STEEL', 'SS316'),
    (r'304L\s+STAINLESS\s+STEEL', 'SS304L'),
    (r'304\s+STAINLESS\s+STEEL', 'SS304'),
    (r'AISI\s*316L', 'SS316L'),
    (r'AISI\s*316', 'SS316'),
    (r'AISI\s*304L', 'SS304L'),
    (r'AISI\s*304', 'SS304'),
    (r'TP\s*316L', 'SS316L'),
    (r'TP\s*316', 'SS316'),
    (r'TP\s*304L', 'SS304L'),
    (r'TP\s*304', 'SS304'),
    (r'SS\s*317L', 'SS317L'),
    (r'SS\s*317', 'SS317'),
    (r'SS\s*316L', 'SS316L'),
    (r'SS\s*316', 'SS316'),
    (r'SS\s*304L', 'SS304L'),
    (r'SS\s*304', 'SS304'),
    (r'\b316L\b', 'SS316L'),
    (r'\b316\b', 'SS316'),
    (r'\b304L\b', 'SS304L'),
    (r'\b304\b', 'SS304'),
    (r'UNS\s*N08825', 'UNS N08825'),
    (r'INCOLOY\s*825|INCOLY\s*825|INCOLLOY\s*825|INC\.?\s*825|ALLOY\s*825', 'ALLOY 825'),
    (r'UNS\s*N06625', 'UNS N06625'),
    (r'INCONEL\s*625|ALLOY\s*625', 'INCONEL 625'),
    (r'HASTELLOY\s*C[-\s]*276', 'HASTELLOY C276'),
    (r'DSS\s+UNS\s*S\s*32205|UNS\s*S\s*32205|DUPLEX\s+S32205', 'UNS S32205'),
    (r'DSS\s+UNS\s*S\s*31803|UNS\s*S\s*31803|DUPLEX', 'UNS S31803'),
    (r'SUPER\s+DUPLEX|UNS\s*S\s*32750', 'UNS S32750'),
    (r'UNS\s*S\s*32760', 'UNS S32760'),
    (r'CARBON\s+STEEL|C\.?\s*S\.?|MILD\s+STEEL|M\.?\s*S\.?', 'CS'),
    (r'SOFT\s+IRON', 'SOFT IRON'),
    (r'\bCOPPER\b', 'COPPER'),
    (r'\bLTCS\b', 'LTCS'),
    (r'\bSS\b', 'SS'),
]

_SW_FILLER_ALIASES: list[tuple[str, str]] = [
    (r'EXFOLIATED\s+EXPANDED\s+GRAPHITE', 'EXFOLIATED EXPANDED GRAPHITE'),
    (r'EXPANDED\s+GRAPHITE', 'EXPANDED GRAPHITE'),
    # GGPL preserves the FLEXIBLE qualifier when the customer states it
    (r'FLEXIBLE\s+GRAPHITE|FLEX\.?\s+GRAPHITE|W/\s*FLEXIBLE\s+GRAPHITE', 'FLEXIBLE GRAPHITE'),
    # EPC gasket lists (Toyo etc.) call the filler "Graphite Tape" (typo "Grapite")
    (r'GRAP[HI]?ITE\s+TAPE', 'GRAPHITE'),
    (r'GRAFOIL|GRAFIL|\bGR?PH\b|GRAPH(?:ITE|OIL)', 'GRAPHITE'),
    # "Vermiculite Fill or amorphous polysilicic acid fibers with talc filler"
    (r'VERMICULITE(?:\s+FILL)?|AMORPHOUS\s+POLYSILICIC', 'VERMICULITE'),
    (r'\bPTFE\b|TEFLON', 'PTFE'),
    (r'\bMICA\b', 'MICA'),
    (r'\bCERAMIC\b', 'CERAMIC'),
]

_SW_MATERIAL_RE = '|'.join(f'(?:{pattern})' for pattern, _ in _SW_MATERIAL_ALIASES)
_SW_FILLER_RE = '|'.join(f'(?:{pattern})' for pattern, _ in _SW_FILLER_ALIASES)


def _sw_prepare_text(description: str) -> str:
    text = description.upper()
    text = text.replace('\xa0', ' ')
    typo_replacements = {
        r'\b(?:SPRL|SPRIAL|SPRIRAL|SPIRIAL|SPLRAL|SPRLAL|SPIRRAL|SPRRAL|SPIR|SPL|SP)\s*[-\s]*(?:W(?:OU)?ND\w*|WIND\w*)\b': 'SPIRAL WOUND',
        r'\bSPIRALWOUND?\b': 'SPIRAL WOUND',
        r'\b(?:WNDLNG|WNDNG|WLDNG|WLNDNG|WINDLNG|WINDNG|WDG)\b': 'WINDING',
        r'\b(?:RLNG|RFNG|RNG|RINGS)\b': 'RING',
        r'\b(?:CENTERLNG|CENTRLNG)\b': 'CENTERING',
        r'\bCENTRING\b': 'CENTERING',
        r'\bSPW(?:D|G)?\b': 'SPIRAL WOUND',
        r'\bSWG\b': 'SPIRAL WOUND GASKET',
        r'\bGRAPH(?:L|I)?TLE\b': 'GRAPHITE',
        r'\b(?:GPH|GRAPH|GRAP)\b': 'GRAPHITE',
        r'\bFLR\b': 'FILLER',
        r'\bINR\b': 'INNER',
        r'\bINER\b': 'INNER',
        r'\bOTR\b': 'OUTER',
        r':-': ':',
        r'\bFLD\b': 'FILLED',
        r'\bSULT\b': 'SUIT',
        # "OUT=CS" / "IN=SS316" compact export notation
        r'\bOUT\s*=\s*': 'OUTER RING ',
        r'\bIN\s*=\s*': 'INNER RING ',
        # "4.5T" compact thickness notation ("4.5T," → "4.5MM THK")
        r'\b(\d+(?:\.\d+)?)T\b(?=[\s,;])': r'\1MM THK',
    }
    for pattern, replacement in typo_replacements.items():
        text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)
    text = re.sub(r'\bS[GC]\s*(304L?|316L?|317L?)\b', r'SS\1', text, flags=re.IGNORECASE)
    text = re.sub(r'(?i)(GASKET)(?=(SKAG|CAM|KAMM|DOUBLE|COPPER|\d))', r'\1 ', text)
    text = re.sub(r'(?i)(SKAG|KAMMPROFILE|CAMPROFILE|PROFILE)(?=(WITH|FOR|OD|ID|\d))', r'\1 ', text)
    text = re.sub(r'(?i)(THK)(?=(CL|CLASS|\d))', r'\1 ', text)
    text = re.sub(r'(?i)(CL\.?\s*\d{2,4})(?=(FOR|PRO|GASKET|SS|TP|AISI))', r'\1 ', text)
    # Customer exports often lose spaces at boundaries such as FillerAlloy or IRAlloy.
    text = re.sub(r'(?i)(FILLER|FILLED|WINDINGS?|WOUND|CL\d{2,4}|IR|OR)(?=(CL\d{2,4}|ALLOY|INCOLOY|INCONEL|SS|STAINLESS|AISI|TP|CS|CARBON|DSS|UNS))', r'\1 ', text)
    text = re.sub(r'(?i)(GRAPHITE|GRAFOIL|PTFE)(?=(FILLER|FILLED))', r'\1 ', text)
    text = re.sub(r'(?i)(\d)(IN\b)', r'\1 IN', text)
    text = re.sub(r'(?i)(\d)(INST\.?\s+KIT|INSULATING|ISOLATING)', r'\1 \2', text)
    text = re.sub(r'(?i)(EN\s*1514)(\d+)(PN\s*\d+)', r'\1 \2 \3', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _sw_norm_material(raw: str | None) -> str | None:
    if not raw:
        return None
    value = re.sub(r'\s+', ' ', raw.strip().upper())
    value = value.strip(' ,:+-/')
    for pattern, canonical in _SW_MATERIAL_ALIASES:
        if re.fullmatch(pattern, value, re.IGNORECASE):
            return canonical
    for pattern, canonical in _SW_MATERIAL_ALIASES:
        if re.search(r'\b' + pattern + r'\b', value, re.IGNORECASE):
            return canonical
    return value or None


def _sw_norm_filler(raw: str | None) -> str | None:
    if not raw:
        return None
    value = re.sub(r'\s+', ' ', raw.strip().upper())
    for pattern, canonical in _SW_FILLER_ALIASES:
        if re.search(pattern, value, re.IGNORECASE):
            return canonical
    return value or None


def _first_match_material(patterns: list[str], text: str) -> str | None:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return _sw_norm_material(match.group('mat'))
    return None


def _first_match_filler(patterns: list[str], text: str) -> str | None:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return _sw_norm_filler(match.group('mat'))
    return None


def _extract_spw_components(description: str) -> dict:
    """Extract SPW construction from broad customer wording.

    The extractor is intentionally evidence-driven: values are only populated
    when they appear near winding/filler/IR/OR/inner/outer ring cues.
    """
    text = _sw_prepare_text(description)
    result: dict = {}

    size = _extract_first_size(text)
    if size:
        result['size'] = size
        result['size_type'] = 'NPS'

    winding = _first_match_material([
        rf'(?:SPIRAL|WINDINGS?|WOUND)\s*[:=-]\s*(?P<mat>{_SW_MATERIAL_RE})',
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:WINDINGS?|WOUND\b)',
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+SPIRAL\s+WOUND\b',
        rf'(?:WINDINGS?|WOUND|MOC\s*:?)\s+(?P<mat>{_SW_MATERIAL_RE})',
        rf'SPIRAL\s+WOUND\s+(?P<mat>{_SW_MATERIAL_RE})',
    ], text)
    if winding:
        result['sw_winding_material'] = winding

    filler = _first_match_filler([
        rf'(?:FILLER|FILLED|FILL)\s*[:=-]\s*(?P<mat>{_SW_FILLER_RE})',
        rf'(?P<mat>{_SW_FILLER_RE})\s+(?:FILLER|FILLED|FILL\b)',
        rf'(?:FILLER|FILLED|FILL)\s+(?P<mat>{_SW_FILLER_RE})',
    ], text)
    if filler:
        result['sw_filler'] = filler

    inner = _first_match_material([
        rf'(?P<mat>{_SW_MATERIAL_RE})[\s-]+(?:INNER\s+RING|I\.?R\.?\b|IR\b)',
        rf'(?:INNER\s+RING|I\.?R\.?\b|IR\b)\s*[-:]*\s*(?P<mat>{_SW_MATERIAL_RE})',
    ], text)
    outer = _first_match_material([
        rf'(?:OUTER\s+RING|CENTER(?:ING)?\s+RING|CENTRE\s+RING|O\.?R\.?\b|OR\b)\s*[-:]*\s*(?P<mat>{_SW_MATERIAL_RE})',
        rf'(?P<mat>{_SW_MATERIAL_RE})[\s-]+(?:OUTER\s+RING|CENTER(?:ING)?\s+RING|CENTRE\s+RING|O\.?R\.?\b|OR\b)',
        rf'(?P<mat>{_SW_MATERIAL_RE})[\s-]+(?:CENTERING|CENTRE|CENTER)\b',
    ], text)

    same_ring = re.search(
        rf'(?P<mat>{_SW_MATERIAL_RE})\s*(?:INNER|OUTER)\s*(?:AND|&)\s*(?:INNER|OUTER|CENTER(?:ING)?|CENTRE)(?:\s+(?:CENTER(?:ING)?\s+)?RING)?'
        rf'|(?P<mat2>{_SW_MATERIAL_RE})\s+(?:INNER|OUTER)\s+RING\s*(?:AND|&)\s*(?:INNER|OUTER)\s+RING',
        text,
        re.IGNORECASE,
    )
    if same_ring:
        both = _sw_norm_material(same_ring.group('mat') or same_ring.group('mat2'))
        inner = inner or both
        outer = outer or both

    if (not inner and winding
            and re.search(r'\b(?:WITH\s+)?I\s+RING\b|\bINNER\s+RING\b', text, re.IGNORECASE)
            and not re.search(r'(?:WITHOUT|W/?O\.?|NO)\s+(?:AN\s+)?INNER\s+RING', text, re.IGNORECASE)):
        inner = winding
    if not outer:
        proceed_outer = re.search(rf'OUTER\s+RING\s+AS\s+"?(?P<mat>{_SW_MATERIAL_RE})"?', text, re.IGNORECASE)
        if proceed_outer:
            outer = _sw_norm_material(proceed_outer.group('mat'))

    # Domain correction for compact/garbled rows: alloy winding SPW typically
    # uses matching alloy IR and CS OR. Some exports concatenate text as
    # "CS IRAlloy 825 OR"; use the construction pattern rather than the bad
    # boundary when both materials are present after the filler.
    if winding and winding not in ('CS', 'LTCS', 'SS') and re.search(rf'\bCS\s+(?:IR|I\.R\.)\s+(?:{_SW_MATERIAL_RE})\s+(?:OR|O\.R\.)\b', text, re.IGNORECASE):
        inner = winding
        outer = 'CS'

    # "CR/IR 316L" — centering + inner ring share one material
    both_after = re.search(
        rf'\b(?:CR|OR)\s*/\s*IR\s*[-:]*\s*(?P<mat>{_SW_MATERIAL_RE})',
        text, re.IGNORECASE,
    )
    if both_after:
        both = _sw_norm_material(both_after.group('mat'))
        inner = inner or both
        outer = outer or both

    # Compact quad notation "SS316/SS316/FG/CS" = winding/inner/filler/outer
    quad = re.search(
        rf'(?P<w>{_SW_MATERIAL_RE})\s*/\s*(?P<i>{_SW_MATERIAL_RE})\s*/\s*(?P<f>FG|GRAPH\w*|PTFE|MICA)\s*/\s*(?P<o>{_SW_MATERIAL_RE})',
        text, re.IGNORECASE,
    )
    if quad:
        winding = winding or _sw_norm_material(quad.group('w'))
        inner = inner or _sw_norm_material(quad.group('i'))
        outer = outer or _sw_norm_material(quad.group('o'))
        if not result.get('sw_filler'):
            # compact "FG" code is quoted as plain GRAPHITE
            f = quad.group('f').upper()
            result['sw_filler'] = 'GRAPHITE' if f in ('FG', 'GRAPH') else _sw_norm_filler(f)
        result['sw_winding_material'] = winding

    # Winding fallback: many enquiries state the alloy once without a WINDING/
    # WOUND cue (e.g. "GASKET SPIRAL WOUND 4.5MM THK; 1"; 300#; AISI 316,
    # GRAPHITE, CS CENTERING / SS INNER RING"). Take the first material token
    # that is not ring/centering context and is not a generic 'SS'/'CS'.
    if not winding:
        ring_kw = r'(?:INNER|OUTER|CENTER(?:ING)?|CENTRE|I\.?R\.?|O\.?R\.?|RING)'
        for m in re.finditer(_SW_MATERIAL_RE, text, re.IGNORECASE):
            before = text[max(0, m.start() - 16):m.start()].upper()
            after = text[m.end():m.end() + 16].upper()
            # skip tokens that belong to a ring spec ("CS CENTERING RING",
            # "INNER RING SS316", "IR SS-316") — ring keyword directly adjacent
            if re.search(rf'\b{ring_kw}\s*[-:=/]*\s*$', before):
                continue
            if re.match(rf'\s*[-:=/]*\s*{ring_kw}\b', after):
                continue
            candidate = _sw_norm_material(m.group(0))
            if candidate in ('SS', 'CS', 'LTCS', None):
                continue
            winding = candidate
            result['sw_winding_material'] = winding
            break

    # RULE Z Part 1 / worked example C — ERP style codes carry the grade inline:
    # "TYPE SPV2F316L/GRAPH" is an SS316L winding with a graphite filler. The
    # grade is glued to the code, so no material-token regex above reaches it.
    if not winding:
        erp = re.search(
            r'\b(?:TYPE\s*)?(?:SPV\d*[A-Z]*|CGI|CG|RWI|RW|WRI|WR|913M?)\s*'
            r'(3\d{2}L?H?)\b',
            text, re.IGNORECASE,
        )
        if erp:
            winding = f'SS{erp.group(1).upper()}'
            result['sw_winding_material'] = winding

    if inner:
        result['sw_inner_ring'] = inner
    if outer:
        result['sw_outer_ring'] = outer

    rating = _rating_from_text(text)
    if rating:
        result['rating'] = rating

    thk_match = re.search(r'(?<!/)(\d+(?:\.\d+)?)\s*(?:MM)?\s*(?:NOM\s+)?(?:GASKET\s+)?THK|(?<!/)(\d+(?:\.\d+)?)\s*MM\s+NOM\s+THK', text, re.IGNORECASE)
    if thk_match:
        result['thickness_mm'] = float(thk_match.group(1) or thk_match.group(2))

    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    return result


def _extract_first_size(text: str) -> str | None:
    s = text.upper()
    _Q = r'["\x94“”″˝]|\'{1,2}'
    match = re.search(rf'\b(?:NPS|SIZE\s+IN\s+INCH|SIZE)\s*:?\s*(\d+(?:\.\d+)?|\d+[\s-]\d+/\d+|\d+/\d+|[¼½¾])\s*(?:{_Q}|INCH|IN)?\b', s)
    if match:
        return _size_from_text(match.group(1))
    match = re.search(rf'\b(\d+(?:\.\d+)?|\d+[\s-]\d+/\d+|\d+/\d+|[¼½¾])\s*(?:{_Q}|INCH|IN)', s)
    if match:
        return _size_from_text(match.group(1))
    # DN sizes keep their DN identity — GGPL prints them as e.g. '25 DN'
    match = re.search(r'\bDN\s*[:\s]?\s*(\d{1,4})\b', s)
    if match:
        return f'DN {match.group(1)}'
    match = re.search(r'\b(\d+(?:\.\d+)?)\s*NB\b', s)
    if match:
        return f'{match.group(1)} NB'
    match = re.search(r'^\s*(\d+(?:\.\d+)?|\d+/\d+|[¼½¾])\s*(?:,|GASKET|INSULATING|ISOLATING|INST\.?|IN\b)', s)
    if match:
        return _size_from_text(match.group(1))
    # "GASKET, 150 mm, ..." — a bare mm value right after the noun is an NB size
    match = re.search(r'\b(?:GASKET|GKT|GSKT)\s*,\s*(\d{2,4})\s*MM\b', s)
    if match:
        return f'{match.group(1)} NB'
    # Concatenated DN/PN exports: "...EN 1514 25PN16..." means DN25 PN16.
    match = re.search(r'\bEN\s*1514\s*(\d{2,4})\s*PN\s*\d{1,2}(?!\d)', s)
    if match:
        return f'{match.group(1)}MM'
    return None


def _parse_number(value: str) -> float:
    return float(value.replace(',', '.'))


def _extract_od_id_thk(text: str) -> tuple[float | None, float | None, float | None, bool]:
    """Return OD, ID, THK, and whether the source was explicitly ID-first."""
    s = text.upper().replace(',', '.')
    od = id_ = thk = None
    id_first = False

    labeled = [
        (r'\bO\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bI\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)', False),
        (r'\bI\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bO\.?D\.?\s*[:=]?\s*(\d+(?:\.\d+)?)', True),
        (r'\bO/?D\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bI/?D\s*[:=]?\s*(\d+(?:\.\d+)?)', False),
        (r'\bI/?D\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b.*?\bO/?D\s*[:=]?\s*(\d+(?:\.\d+)?)', True),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O\.?D\.?\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I\.?D\.?\b', False),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I\.?D\.?\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O\.?D\.?\b', True),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O/?D\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I/?D\b', False),
        (r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+I/?D\b.*?\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O/?D\b', True),
        (r'\bINSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM.*?\bOUTSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM', True),
        (r'\bOUTSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM.*?\bINSIDE\s+DIAMETER\s*=?\s*(\d+(?:\.\d+)?)\s*MM', False),
    ]
    for pattern, reversed_order in labeled:
        match = re.search(pattern, s, re.IGNORECASE)
        if match:
            first, second = _parse_number(match.group(1)), _parse_number(match.group(2))
            if reversed_order:
                id_, od = first, second
                id_first = True
            else:
                od, id_ = first, second
            break

    if od is None or id_ is None:
        # Common custom gasket shorthand: ID x OD x THK.
        match = re.search(r'(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?', s)
        if match:
            first, second, third = map(_parse_number, match.groups())
            id_, od, thk = first, second, third
            id_first = True

    if od is None or id_ is None:
        match = re.search(r'\bOD\s*(\d+(?:\.\d+)?)\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*ID\s*(\d+(?:\.\d+)?)', s)
        if match:
            od, thk, id_ = map(_parse_number, match.groups())

    if od is None or id_ is None:
        match = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+OD\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s+ID\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s*THK', s)
        if match:
            od, id_, thk = map(_parse_number, match.groups())

    if od is None or id_ is None:
        match = re.search(r'\b(\d+(?:\.\d+)?)\s*(?:MM)?\s+O/?D\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s+I/?D(?:\s*[xX]\s*(\d+(?:\.\d+)?)\s*(?:MM)?\s*THK)?', s)
        if match:
            od, id_ = _parse_number(match.group(1)), _parse_number(match.group(2))
            if match.group(3):
                thk = _parse_number(match.group(3))

    thk_match = re.search(r'\b(?:THK|THICKNESS)\s*[:=]?\s*(\d+(?:\.\d+)?)\s*(?:MM)?\b|(\d+(?:\.\d+)?)\s*(?:MM)?\s*(?:THK|THICK)\b', s)
    if thk_match:
        thk = _parse_number(thk_match.group(1) or thk_match.group(2))
    else:
        # Pattern like "OD 1430 x3x ID 1404".
        mid = re.search(r'\bOD\s*\d+(?:\.\d+)?\s*[xX]\s*(\d+(?:\.\d+)?)\s*[xX]\s*ID\b', s)
        if mid:
            thk = _parse_number(mid.group(1))

    return od, id_, thk, id_first


def _extract_rtj_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    size = _extract_first_size(text)
    if size:
        result['size'] = size
        result['size_type'] = 'NPS'
    rating = _rating_from_text(text)
    if rating:
        result['rating'] = rating
    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    ring = re.search(r'\b(?P<ring>R|RX|BX)\s*[- ]?\s*(?P<num>\d{1,4})\b', text, re.IGNORECASE)
    if ring:
        result['ring_no'] = f'{ring.group("ring").upper()}-{ring.group("num")}'

    if re.search(r'\bOCT(?:AGONAL)?\b|TYPE\s*O\b|8[-\s]*SIDED', text, re.IGNORECASE):
        result['rtj_groove_type'] = 'OCTAGONAL'
    elif re.search(r'\bOVAL\b|ELLIPTICAL|TYPE\s*R\b', text, re.IGNORECASE):
        result['rtj_groove_type'] = 'OVAL'
    elif re.search(r'\bBX\b', text, re.IGNORECASE):
        result['rtj_groove_type'] = 'BX'

    material = None
    for pattern in (
        r'(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)\s+(?:OCTAGONAL|OVAL|RING\s+JOINT|RING\s+TYPE|RTJ|R/?J)',
        r'(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)\s+(?:OCTAGONAL|OVAL)\s+RING\s+GASKETS?',
        r'(?:OCTA\s+)?R/?J\s+\d{2,4}#?\s*,?\s*(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)',
        r'(?:MOC|MATERIAL)\s*:?\s*(?P<mat>SOFT\s+IRON|INCOLOY\s*825|INCOLY\s*825|ALLOY\s*825|INCONEL\s*625|UNS\s*S\s*3\d{4}|SS[-\s]*316L?|316L?SS|SS[-\s]*304L?|304L?SS|F\d{1,2}|LOW\s+CARBON\s+STEEL|LTCS|MONEL\s*400|HASTELLOY\s*C[-\s]*276)',
    ):
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            material = _sw_norm_material(match.group('mat'))
            break
    if material:
        if material == 'SOFT IRON':
            material = 'SOFTIRON'
        if re.search(r'GALVANI[ZS]ED|ZINC\s+PLATED', text, re.IGNORECASE) and 'GALVANISED' not in material:
            material += ' GALVANISED'
        result['moc'] = material
    else:
        fallback = _first_match_material([rf'(?P<mat>{_SW_MATERIAL_RE})'], text)
        if fallback:
            result['moc'] = 'SOFTIRON' if fallback == 'SOFT IRON' else fallback

    hardness = re.search(r'\b(\d{2,3})\s*(?:BHN|HB|HRB|HRBW)\b', text, re.IGNORECASE)
    if hardness:
        result['rtj_hardness_bhn'] = float(hardness.group(1))
    return result


def _extract_kamm_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    od, id_, thk, _ = _extract_od_id_thk(text)
    if od is not None and id_ is not None:
        result.update({'size_type': 'OD_ID', 'od_mm': od, 'id_mm': id_})
    else:
        size = _extract_first_size(text)
        if size:
            result['size'] = size
            result['size_type'] = 'NPS'
    rating = _rating_from_text(text)
    if rating:
        result['rating'] = rating
    if thk is not None:
        result['thickness_mm'] = thk
    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    # Compact KAMM notation: PRO+FILL+CR: SS316L+GRAPHITE+SS316L
    compact = re.search(r'PRO\s*\+\s*FILL\s*\+\s*(?:CR|CENTER(?:ING)?\s*RING)\s*:?\s*(?P<core>[^+,\s]+(?:\s*316L?)?)\s*\+\s*(?P<surface>[^+,\s]+)(?:\s*\([^)]*\))?\s*\+\s*(?P<ring>[^,\s]+(?:\s*316L?)?)', text, re.IGNORECASE)
    if compact:
        result['kamm_core_material'] = _sw_norm_material(compact.group('core'))
        result['kamm_surface_material'] = _sw_norm_filler(compact.group('surface')) or _sw_norm_material(compact.group('surface'))
        result['sw_outer_ring'] = _sw_norm_material(compact.group('ring'))

    if not result.get('kamm_core_material'):
        core = _first_match_material([
            rf'(?:PROFILE|CAM\s*PROFILE|KAMM?PROFILE|GROOVED\s+PROFILE|GROOVED\s+METAL|CORE|INSERT)\s*(?:MATERIAL)?\s*:?\s*(?P<mat>{_SW_MATERIAL_RE})',
            rf'(?P<mat>{_SW_MATERIAL_RE})\s*(?:KAMM?PROFILE|CAM\s*PROFILE|GROOVED\s+PROFILE|PROFILE\s+GASKET)',
            rf'(?:KAMM?PROFILE|CAM\s*PROFILE|GROOVED\s+PROFILE|GMGC).*?(?P<mat>{_SW_MATERIAL_RE})',
            rf'WITH\s+(?P<mat>{_SW_MATERIAL_RE})\s+AND\s+(?:{_SW_MATERIAL_RE})\s+CENTER(?:ING)?\s+RING',
            rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:GPH|GRAPHITE|PTFE)',
            rf'\b(?P<mat>{_SW_MATERIAL_RE})\s*/\s*(?:{_SW_FILLER_RE})',
        ], text)
        if core:
            result['kamm_core_material'] = core

    if not result.get('kamm_surface_material'):
        surface = _first_match_filler([
            rf'(?:LAYER|LAYERS|FACING|COVERING|FILL(?:ER)?)\s*(?:MATERIAL)?\s*:?\s*(?P<mat>{_SW_FILLER_RE})',
            rf'(?P<mat>{_SW_FILLER_RE})\s*(?:LAYER|LAYERS|FACING|COVERING|FILLER)',
            rf'(?:{_SW_MATERIAL_RE})\s+(?P<mat>GPH|GRAPHITE|PTFE)',
            rf'(?:{_SW_MATERIAL_RE})\s*/\s*(?P<mat>{_SW_FILLER_RE})',
        ], text)
        if surface:
            result['kamm_surface_material'] = surface

    if not result.get('sw_outer_ring'):
        outer = _first_match_material([
            rf'(?:CENTER(?:ING)?\s+RING|CENTRING\s+RING|OUTER\s+RING|CR)\s*:?\s*(?P<mat>{_SW_MATERIAL_RE})',
            rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:CENTER(?:ING)?\s+RING|CENTRING\s+RING|OUTER\s+RING)',
            rf'\b(?:INR\s+)?(?:{_SW_MATERIAL_RE})\s+(?P<mat>{_SW_MATERIAL_RE})\s+CENTER(?:ING)?\s+RING',
        ], text)
        if outer:
            result['sw_outer_ring'] = outer

    if not result.get('size') and result.get('size_type') != 'OD_ID':
        trailing_size = re.search(r',\s*(\d+(?:\.\d+)?)\s*$', text)
        if trailing_size:
            result['size'] = _size_from_text(trailing_size.group(1))
            result['size_type'] = 'NPS'

    core_thk = re.search(r'\bCORE\s+THK\s*[:=]?\s*(\d+(?:\.\d+)?)\s*MM\b|\((\d+(?:\.\d+)?)\s*MM\s+CORE\s+THK\)', text, re.IGNORECASE)
    if core_thk:
        result['kamm_core_thk'] = float(core_thk.group(1) or core_thk.group(2))
    # A core thickness is never guessed here. RULE Y Part 4 derives it from the
    # total (core = total - 2 x layer) and checks it against GGPL stock; the old
    # flat 4.0MM fallback fabricated a core unrelated to the stated total.

    if re.search(r'\bINTEGRAL\s+(?:OUTER\s+)?RING\b', text, re.IGNORECASE):
        result['kamm_integral_outer_ring'] = True

    if result.get('kamm_core_material'):
        result['sw_winding_material'] = result['kamm_core_material']
    if result.get('kamm_surface_material'):
        result['sw_filler'] = result['kamm_surface_material']
    return result


def _extract_dji_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    od, id_, thk, id_first = _extract_od_id_thk(text)
    if od is not None and id_ is not None:
        result.update({'size_type': 'OD_ID', 'od_mm': od, 'id_mm': id_, 'dji_id_first': id_first})
    if thk is not None:
        result['thickness_mm'] = thk

    jacket = _first_match_material([
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+(?:DOUBLE\s+)?JACKET(?:ED)?',
        rf'(?:JACKET(?:ED)?|MATL\.?|MATERIAL)\s*(?:MATERIAL)?\s*:?\s*(?P<mat>{_SW_MATERIAL_RE})',
        rf'MATERIAL\s+(?P<mat>{_SW_MATERIAL_RE})\s+AND',
    ], text)
    if jacket:
        result['moc'] = jacket

    filler = _first_match_filler([
        rf'(?:AND|WITH)\s+(?P<mat>{_SW_FILLER_RE})\b',
        rf'(?P<mat>{_SW_FILLER_RE})\s+FILLER',
    ], text)
    if filler:
        result['dji_filler'] = filler

    if re.search(r'\bDRAWING\b|AS\s+PER\s+DRAWING', text, re.IGNORECASE):
        result['special'] = 'AS PER DRAWING'
    if re.search(r'\bRF\b|RAISED\s+FACE', text, re.IGNORECASE):
        result['dji_face_type'] = 'RF'
    elif re.search(r'\bFF\b|FULL\s+FACE', text, re.IGNORECASE):
        result['dji_face_type'] = 'FF'
    return result


def _extract_isk_components(description: str) -> dict:
    text = _sw_prepare_text(description)
    result: dict = {}
    compact_size_rating = re.search(r'^\s*(?P<size>\d{1,2}(?:\.\d+)?)(?P<rating>150|300|600|900|1500|2500)#', text)
    if compact_size_rating:
        result['size'] = _size_from_text(compact_size_rating.group('size'))
        result['size_type'] = 'NPS'
        result['rating'] = f'{compact_size_rating.group("rating")}#'
    size = _extract_first_size(text)
    if size and not result.get('size'):
        result['size'] = size
        result['size_type'] = 'NPS'
    rating = _rating_from_text(text)
    if rating and not result.get('rating'):
        result['rating'] = rating
    standard = _standard_from_text(text)
    if standard:
        result['standard'] = standard

    if re.search(r'\bTYPE[-\s]?D\b|\bRTJ\b|\bR/?J\b', text, re.IGNORECASE):
        result['gasket_type'] = 'ISK_RTJ'
        result['isk_style'] = 'STYLE-N'
        result['isk_type'] = 'TYPE-D'
    elif re.search(r'\bSTYLE[-\s]?CS\b|\bVCS\b|COMMANDER\s+EXTREME', text, re.IGNORECASE):
        result['isk_style'] = 'STYLE-CS'
        result['isk_type'] = 'TYPE-F'
    elif re.search(r'\bTYPE[-\s]?E\b', text, re.IGNORECASE):
        result['isk_style'] = 'TYPE-E'
        result['isk_type'] = 'TYPE-E'
    elif re.search(r'\bTYPE[-\s]?F\b', text, re.IGNORECASE):
        result['isk_style'] = 'TYPE-F'
        result['isk_type'] = 'TYPE-F'
    elif re.search(r'\bSTYLE[-\s]?N\b', text, re.IGNORECASE):
        result['isk_style'] = 'STYLE-N'

    if re.search(r'\bFF\b|FULL\s+FACE|TYPE[-\s]?E\b', text, re.IGNORECASE):
        result['face_type'] = 'FF'
    elif re.search(r'\bRF\b|RAISED\s+FACE|TYPE[-\s]?F\b|RTJ\b', text, re.IGNORECASE):
        result['face_type'] = 'RF'

    gasket_mat = re.search(r'\b(?:GRE|G[-\s]?10|G[-\s]?11|GLASS\s+REINFORCED\s+EPOXY)[A-Z0-9\s()/-]*', text, re.IGNORECASE)
    if gasket_mat:
        mat = gasket_mat.group(0).strip()
        grade = re.search(r'G[-\s]?(10|11)', mat, re.IGNORECASE)
        result['isk_gasket_material'] = f'GRE G-{grade.group(1)}' if grade else 'GRE'
        result['isk_sleeve_material'] = result['isk_gasket_material']
        result['isk_insulating_washer'] = result['isk_gasket_material']

    core = _first_match_material([
        rf'(?:W/|WITH)?\s*(?P<mat>{_SW_MATERIAL_RE})\s+(?:STEEL\s+)?CORE',
        rf'(?P<mat>{_SW_MATERIAL_RE})\s+STEEL\s+CORE',
    ], text)
    if core:
        result['isk_core_material'] = core

    washer = _first_match_material([
        rf'(?P<mat>ZINC\s+PLATED\s+CS|{_SW_MATERIAL_RE})\s+WASHER',
        rf'METALLIC\s+WASHER\s+(?P<mat>ZINC\s+PLATED\s+CS|{_SW_MATERIAL_RE})',
    ], text)
    if washer:
        result['isk_washer_material'] = washer

    if re.search(r'\bPTFE\b|TEFLON', text, re.IGNORECASE):
        result['isk_primary_seal'] = 'PTFE SPRING ENERGISED SEAL' if re.search(r'SPRING|ENERGI[ZS]ED|COMMANDER', text, re.IGNORECASE) else 'PTFE'
    if re.search(r'NON[-\s]?FIRE\s+SAFE|SPRING|ENERGI[ZS]ED', text, re.IGNORECASE):
        result['isk_fire_safety'] = 'NON FIRE SAFE'
    elif re.search(r'\bFIRE\s+SAFE\b', text, re.IGNORECASE):
        result['isk_fire_safety'] = 'FIRE SAFE'

    return result


def _material_from_text(description: str) -> str | None:
    patterns = [
        r'MATERIAL\s+STANDARD\s*:?\s*([^,;]+)',
        r'MATERIAL\s+OF\s+CONSTRUCTION\s*:?\s*([^,;\n]+)',
        r'\bMOC\s*:?\s*([^,;]+)',
        r'\bMATERIAL\s*:?\s*([^,;]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, description, re.IGNORECASE)
        if match:
            value = match.group(1).strip()
            value = re.sub(r'\s+LOCATION\s*:.*$', '', value, flags=re.IGNORECASE).strip()
            return value or None
    # Generic "elastomer" soft cut gaskets are quoted as EPDM; a shore-A
    # hardness range in the enquiry is carried into the GGPL description.
    if re.search(r'\bELASTOM\w*\b', description, re.IGNORECASE):
        shore = re.search(
            r'(?:SHORE\s*A?\s*(?:HARDNESS)?\s*)?(\d{2})\s*[-–]\s*(\d{2})(?:\s*SHORE\s*A?)?',
            description, re.IGNORECASE,
        )
        if shore and re.search(r'SHORE', description, re.IGNORECASE):
            return f'EPDM {shore.group(1)} - {shore.group(2)} SHORE A HARDNESS'
        return 'EPDM'
    # Grade-qualified echoes come first — GGPL keeps the qualifier
    # Cloth-inserted rubber sheet (EPC gasket lists: "NBR ... rubber with Cloth")
    cloth = re.search(r'\b(NBR|EPDM|SILICONE?|NEOPRENE|SBR|BUTYL)\b[^,;]{0,60}?WITH\s+CLOTH', description, re.IGNORECASE)
    if cloth:
        base = cloth.group(1).upper()
        base = 'SILICONE' if base.startswith('SILICON') else base
        return f'{base} WITH CLOTH INSERT'
    if re.search(r'COMP\.?\s*NON[\s.–-]*ASB\.?\s*SYNTHETIC\s+FIBER', description, re.IGNORECASE):
        return 'COMPRESSED NON ASBESTOS SYNTHETIC FIBER'
    match = re.search(r'FLEX\.?(?:IBLE)?\s+GRAPHITE\s+REINFORCED\s+W(?:/|ITH)\s*(SS\s*\d{3}L?)\s+SHEET\s+INSERT', description, re.IGNORECASE)
    if match:
        grade = match.group(1).replace(' ', '').upper()
        return f'FLEXIBLE GRAPHITE REINFORCED W/{grade} SHEET INSERT'
    match = re.search(r'NON[\s–-]*ASBESTOS\s+(BS\s*7531\s+GR(?:ADE)?\.?\s*[A-Z])', description, re.IGNORECASE)
    if match:
        return f'NON-ASBESTOS {re.sub(r"(?i)GRADE", "GR", match.group(1)).upper()}'
    match = re.search(r'GRAPHITE\b.{0,40}?\bWITH\s+(SS\s*\d{3}L?)\s+TANGED\s+INSERT', description, re.IGNORECASE)
    if match:
        grade = match.group(1).replace(' ', '').upper()
        return f'GRAPHITE WITH {grade} TANGED INSERT'
    if re.search(r'KROLL(?:ER)?\s*&\s*ZILLER', description, re.IGNORECASE):
        return 'KROLLER & ZILLER (G-S-T-P/S) WITH SPACER'
    # EPDM to EN 681 (water/sewerage flange service) is quoted with steel insert
    if re.search(r'\bE[PD][DP]M\b', description, re.IGNORECASE) and re.search(r'EN\s*681|SEWERAGE', description, re.IGNORECASE):
        return 'EPDM RUBBER WITH STEEL INSERT'
    material_aliases = [
        (r'\bBUTYL\s+RUBBER\b', 'BUTYL RUBBER'),
        (r'\bNBR\b|\bNITRILE\b', 'NBR'),
        (r'\bCNAF\b|COMPRESSED\s+NON[\s–-]*ASBESTOS', 'CNAF'),
        (r'\bMODIFIED\s+PTFE\b', 'MODIFIED PTFE'),
        (r'\bRPTFE\b', 'RPTFE'),
        (r'\bPTFE\b|\bTEFLON\b', 'PTFE'),
        (r'\bEXPANDED\s+GRAPHITE\b', 'EXPANDED GRAPHITE'),
        (r'\bPURE\s+GRAPHITE\b', 'PURE GRAPHITE'),
        (r'\bGRAFOIL\b', 'GRAFOIL'),
        (r'ARAMIDIC\s+FIBER', 'ARAMIDIC FIBER'),
        (r'\bARAFBR\b', 'ARAMID FIBER'),
        (r'\bTEMASIL\b|NON[\s–-]*ASBESTOS', 'NON ASBESTOS'),
        (r'\bEPDM\b|\bEDPM\b', 'EPDM'),
        (r'\bVITON\b', 'VITON'),
        (r'\bNATURAL\s+RUBBER\b|\bNAT\.?\s*RUB\b', 'NATURAL RUBBER'),
        (r'\bNEOPRENE\b', 'NEOPRENE'),
        (r'\bSILICON(?:E)?\s+RUBBER\b', 'SILICONE'),
    ]
    for pattern, canonical in material_aliases:
        if re.search(pattern, description, re.IGNORECASE):
            return canonical
    return None


def _pre_clean_description(desc: str) -> str:
    """Split concatenated tokens produced by lossy Excel/ERP exports.

    e.g. "Restriction OrificeDN25 CLASS150STAINLESS STEEL 316ASME B16.5"
         "GASKET FLAT - THK=2mmDN 4;CL150RFTEMASIL HTASME B16.21"
         "4300#SPIRAL WOUND GASKET AISI 316..."
    """
    t = _collapse_spaced_abbrevs(desc)  # RULE J-2: "S P W D" → "SPWD" before token fixes
    t = re.sub(r'(?i)(?<=[a-z0-9])(DN\s*\d)', r' \1', t)
    t = re.sub(r'(?i)(DN\s*\d{1,4})(?=[A-Z])', r'\1 ', t)
    t = re.sub(r'(?i)(NPS)(\d)', r'\1 \2', t)
    t = re.sub(r'(?i)^(\d{1,2}(?:\.\d+)?)(CL\.?\s*\d{2,4})', r'\1" \2', t)
    t = re.sub(r'(?i)(\d)(GASKET)', r'\1 \2', t)
    t = re.sub(r'(?i)\b(150|300|400|600|900|1500|2500)(RF|FF)\b', r'\1# \2', t)
    t = re.sub(r'(?i)_(?=\d)', ' ', t)
    t = re.sub(r'(?i)(\d)\s*X\s*(PN\s*\d)', r'\1 X \2', t)
    t = re.sub(r'(?i)(\d)X(PN|DN)', r'\1 X \2', t)
    t = re.sub(r'(?i)GASKETS(?=SPIRAL|RTJ|SW\b)', 'GASKET ', t)
    t = re.sub(r'(?i)(RTJ)(OCT)', r'\1 \2', t)
    t = re.sub(r'(?i)(GASKET)(RF|FF)\b', r'\1 \2', t)
    t = re.sub(r'(?i)(CL(?:ASS)?\.?\s*\d{2,4})(?=[A-Z])', r'\1 ', t)
    t = re.sub(r'(?i)\b(RF|FF)(?=[A-Z]{3})', r'\1 ', t)
    t = re.sub(r'(#)(?=[A-Za-z])', r'# ', t)
    t = re.sub(r'(?i)(THK\s*=?\s*\d+(?:\.\d+)?\s*MM)(?=[A-Z])', r'\1 ', t)
    t = re.sub(r'(?i)(\d)(ASME|ANSI|DIN\b|API\b)', r'\1 \2', t)
    t = re.sub(r'(?i)(?<=[A-Z])(ASME|ANSI)(?=\s*B\s*16)', r' \1', t)
    t = re.sub(r'(?i)(\d{3})(SS\b)', r'\1 \2', t)
    t = re.sub(r'(?i)(CS)(GRAPHITE)', r'\1 \2', t)
    t = re.sub(r'(?i)(RING)(\d{3,4})', r'\1 \2', t)
    t = re.sub(r'(?i)\b(IN)(\d{3,4})\b', r'\1 \2', t)
    # "EN 151425PN16" → "EN 1514 DN25 PN16" (EN 1514 sheet + DN size + PN class)
    t = re.sub(r'(?i)(EN\s*1514)[-\s]*(\d{2,4})\s*(PN\s*\d{1,2})(?!\d)', r'\1 DN\2 \3', t)
    # rating+thickness concatenation: "#1503mm" → "# 150 3mm"
    t = re.sub(r'(?i)#\s*(150|300|400|600|900|1500|2500)(\d(?:\.\d+)?)\s*MM', r'# \1 \2MM', t)
    # size+rating concatenation: "4300#" → '4" 300#', "0.5300#" → '0.5" 300#'
    t = re.sub(
        r'(?<![\d.])(\d{1,2}(?:\.\d{1,2})?)(150|300|400|600|900|1500|2500)\s*#',
        lambda m: f'{m.group(1)}" {m.group(2)}#',
        t,
    )
    return t


def _first_metal(text: str) -> str | None:
    """First metal/alloy token in the text that is not ring context."""
    ring_kw = r'(?:INNER|OUTER|CENTER(?:ING)?|CENTRE|RING)'
    for m in re.finditer(_SW_MATERIAL_RE, text, re.IGNORECASE):
        before = text[max(0, m.start() - 14):m.start()].upper()
        after = text[m.end():m.end() + 14].upper()
        if re.search(rf'\b{ring_kw}\s*[-:=/]*\s*$', before) or re.match(rf'\s*[-:=/]*\s*{ring_kw}\b', after):
            continue
        candidate = _sw_norm_material(m.group(0))
        if candidate not in ('SS', None):
            return candidate
    return None


def _enrich_specialty(item: dict, desc: str, upper: str) -> None:
    """Field extraction for the beyond-six specialty families (v3.1 supplement)."""
    gtype = item.get('gasket_type')

    if gtype == 'SHEET':
        m = re.search(
            r'(\d+(?:\.\d+)?)\s*(MTR|M\b|MM)?\s*[X×]\s*(\d+(?:\.\d+)?)\s*(MTR|M\b|MM)?'
            r'(?:\s*[X×]\s*(\d+(?:\.\d+)?)\s*(?:MM)?)?',
            upper,
        )
        if m:
            def _mm(v: str, unit: str | None) -> float:
                val = float(v)
                unit = (unit or '').strip()
                return val * 1000 if unit in ('M', 'MTR') else val
            if (m.group(2) or '').strip() in ('M', 'MTR') or (m.group(4) or '').strip() in ('M', 'MTR'):
                item['sheet_unit'] = 'MTR'
            item.setdefault('sheet_length_mm', _mm(m.group(1), m.group(2)))
            item.setdefault('sheet_width_mm', _mm(m.group(3), m.group(4)))
            if m.group(5) and not item.get('thickness_mm'):
                item['thickness_mm'] = float(m.group(5))
        if 'ROLL' in upper:
            item['sheet_is_roll'] = True
        if not item.get('moc'):
            item['moc'] = _material_from_text(desc) or _first_metal(upper)
        return

    if gtype == 'MANHOLE':
        m = re.search(r'(\d{2,4}(?:\.\d+)?)\s*[X×]\s*(\d{2,4}(?:\.\d+)?)', upper)
        if m:
            item.setdefault('obround_a_mm', float(m.group(1)))
            item.setdefault('obround_b_mm', float(m.group(2)))
        if re.search(r'\bMCR\b', upper):
            item['manhole_style'] = 'MCR'
        elif re.search(r'\bMC\b', upper):
            item['manhole_style'] = 'MC'
        if re.search(r'SPIRAL|\bSPW\b|\bSWG\b', upper):
            comp = _extract_spw_components(desc)
            for key in ('sw_winding_material', 'sw_filler', 'sw_outer_ring'):
                if comp.get(key) and not item.get(key):
                    item[key] = comp[key]
        if not item.get('moc'):
            item['moc'] = _material_from_text(desc) or ('GRAPHITE' if 'GRAPH' in upper else None)
        return

    if gtype == 'ENVELOPE':
        m = re.search(r'WITH\s+([A-Z0-9/ ]{2,30}?)\s+INSERT', upper)
        if m:
            item.setdefault('envelope_insert', m.group(1).strip())
        elif not item.get('envelope_insert'):
            mat = _material_from_text(desc)
            if mat and 'PTFE' not in mat.upper():
                item['envelope_insert'] = mat
        return

    if gtype == 'LENS':
        if not item.get('moc'):
            if '1.4571' in upper:
                item['moc'] = 'SS316TI'
            else:
                item['moc'] = _first_metal(upper) or ('SOFT IRON' if 'SOFT IRON' in upper else None)
        return

    # OD/ID-dimensioned families
    od, id_, thk, _swapped = _extract_od_id_thk(upper)
    if od and id_:
        item.setdefault('od_mm', od)
        item.setdefault('id_mm', id_)
        item['size_type'] = 'OD_ID'
    if thk and not item.get('thickness_mm'):
        item['thickness_mm'] = thk

    if gtype == 'CMG':
        if 'PTFE' in upper and 'FACING' in upper:
            item['cmg_facing'] = 'PTFE'
        if re.search(r'\bPLAIN\b', upper):
            item['cmg_plain'] = True
        if not item.get('moc'):
            item['moc'] = _first_metal(upper)
    elif gtype in ('METAL_CLAD', 'SOLID_METAL', 'PLUG_GASKET'):
        if not item.get('od_mm') or not item.get('id_mm'):
            m = re.search(r'\bOD\s*[:\s]?\s*(\d{1,4}(?:\.\d+)?)\b.{0,20}?\bID\s*[:\s]?\s*(\d{1,4}(?:\.\d+)?)\b', upper)
            if m:
                item['od_mm'] = float(m.group(1))
                item['id_mm'] = float(m.group(2))
                item['size_type'] = 'OD_ID'
        if not item.get('moc'):
            # the gasket itself is a ring here, so no ring-context exclusion
            m_metal = re.search(_SW_MATERIAL_RE, upper, re.IGNORECASE)
            item['moc'] = ('SOFT IRON' if 'SOFT IRON' in upper else None) \
                or (_sw_norm_material(m_metal.group(0)) if m_metal else None) \
                or _material_from_text(desc)
    elif gtype == 'LIP_SEAL':
        if re.search(r'SPRING\s+ENERG', upper):
            item['spring_energised'] = True
        if not item.get('moc'):
            item['moc'] = _material_from_text(desc) or 'PTFE'
    elif gtype == 'DIAPHRAGM':
        if not item.get('od_mm'):
            m = re.search(r'\bOD\s*[:\s]?\s*(\d{2,4}(?:\.\d+)?)', upper)
            if m:
                item['od_mm'] = float(m.group(1))
        if re.search(r'FABRIC|CLOTH|NYLON|POLYESTER', upper):
            item['fabric_reinforced'] = True
        if not item.get('moc'):
            item['moc'] = _material_from_text(desc)
    elif gtype == 'EYELET':
        m = re.search(r'(SS\s*\d{3}L?|COPPER|BRASS)\s+(?:INNER\s+)?EYELET|EYELET\w*\s+(SS\s*\d{3}L?|COPPER|BRASS)', upper)
        if m:
            item['eyelet_material'] = (m.group(1) or m.group(2)).replace(' ', '')
        if not item.get('moc'):
            item['moc'] = _material_from_text(desc)


def _enrich_from_description(item: dict) -> dict:
    desc = item.get('raw_description') or item.get('description') or ''
    if not desc:
        return item
    # Family rule (GGPL ground truth): a bare "<size><class>#SPIRAL WOUND
    # GASKETRF" enquiry with no materials at all is quoted as a graphite
    # sheet gasket with SS316L tanged insert (soft cut, 3MM, B16.21) —
    # this customer's "spiral wound" wording refers to reinforced graphite.
    if re.fullmatch(
        r'\s*\d+(?:\.\d+)?(?:150|300|400|600|900|1500|2500)#\s*SPIRAL\s+WOUND\s+GASKET\s*(?:RF|FF)?\s*',
        desc, re.IGNORECASE,
    ):
        item['gasket_type'] = 'SOFT_CUT'
        item.setdefault('moc', 'GRAPHITE WITH SS316L INSERT')
        item.setdefault('thickness_mm', 3)

    desc = _pre_clean_description(desc)
    upper = desc.upper()

    item.setdefault('gasket_type', _infer_gasket_type(desc) or 'SOFT_CUT')

    if not item.get('rating'):
        rating = _rating_from_text(desc)
        if rating:
            item['rating'] = rating

    if not item.get('thickness_mm'):
        thk_match = re.search(r'(?<!/)(\d+(?:\.\d+)?)\s*(?:MM)?\s*(?:THK|THICK)', upper)
        if thk_match:
            item['thickness_mm'] = float(thk_match.group(1))
        else:
            thk_match = re.search(
                r'\b(\d+(?:\.\d+)?)\s*MM\b(?=\s*,?\s*(?:ASME|ANSI|B\s*16|EN\s*1514|$))',
                upper,
            )
            if thk_match:
                item['thickness_mm'] = float(thk_match.group(1))

    if not item.get('standard'):
        standard = _standard_from_text(desc)
        if standard:
            item['standard'] = standard

    if not item.get('face_type'):
        if re.search(r'\b(?:FULL\s+FACE|FF)\b', upper):
            item['face_type'] = 'FF'
        elif re.search(r'\bRF\b|\bRAISED\s+FACE\b', upper):
            item['face_type'] = 'RF'

    gasket_type = item.get('gasket_type')

    if gasket_type == 'RTJ':
        extracted = _extract_rtj_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value

    if gasket_type == 'KAMM':
        extracted = _extract_kamm_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value
        if any(item.get(k) for k in ('kamm_core_material', 'kamm_surface_material', 'sw_winding_material')):
            item.pop('moc', None)

    if gasket_type == 'DJI':
        extracted = _extract_dji_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value

    if gasket_type in ('ISK', 'ISK_RTJ'):
        extracted = _extract_isk_components(desc)
        for key, value in extracted.items():
            if value and not item.get(key):
                item[key] = value
        if extracted.get('gasket_type'):
            item['gasket_type'] = extracted['gasket_type']
            gasket_type = item['gasket_type']

    if item.get('gasket_type') == 'SPIRAL_WOUND':
        extracted = _extract_spw_components(desc)
        for key, value in extracted.items():
            if value and (not item.get(key) or (key == 'rating' and '/' in str(value))):
                item[key] = value

        # A generic pipe/spec material column sometimes arrives as `moc`
        # beside a rich SPW description. Once component fields are present,
        # the SPW MOC must be rebuilt from those fields, not from that context.
        if item.get('moc') and any(item.get(k) for k in (
            'sw_winding_material', 'sw_filler', 'sw_inner_ring', 'sw_outer_ring'
        )):
            if not re.search(r'\b(?:WIND|FILL|INNER|OUTER|CENTER|CENTRE|IR|OR|SPW|SPIRAL)\b', str(item['moc']), re.IGNORECASE):
                item.pop('moc', None)

        if not item.get('sw_winding_material'):
            match = re.search(
                r'\b(?:WINDING|MOC\s*:?)\s+(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\b'
                r'|\b(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\s+WINDING\b',
                upper,
            )
            if match:
                material = match.group(1) or match.group(2)
                item['sw_winding_material'] = re.sub(r'\s+', '', material.upper()).replace('INCONEL', 'INCONEL ').replace('ALLOY', 'ALLOY ')
        if not item.get('sw_filler'):
            if re.search(r'\b(?:GRAPH(?:ITE|OIL)|GRAFOIL)\b', upper):
                item['sw_filler'] = 'GRAPHITE'
            elif re.search(r'\bPTFE\b', upper):
                item['sw_filler'] = 'PTFE'
        if not item.get('sw_inner_ring'):
            match = re.search(
                r'\bINNER\s+RING\s+(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\b'
                r'|\b(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\s+INNER\s+RING\b',
                upper,
            )
            if match:
                material = match.group(1) or match.group(2)
                item['sw_inner_ring'] = re.sub(r'\s+', '', material.upper()).replace('INCONEL', 'INCONEL ').replace('ALLOY', 'ALLOY ')
        if not item.get('sw_outer_ring'):
            match = re.search(
                r'\b(?:OUTER|CENTERING|CENTRE|CENTER)\s+RING\s+(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\b'
                r'|\b(SS\s*\d{3}L?|CS|LTCS|ALLOY\s*625|INCONEL\s*\d+|HASTELLOY\s*C276)\s+(?:OUTER|CENTERING|CENTRE|CENTER)\s+RING\b',
                upper,
            )
            if match:
                material = match.group(1) or match.group(2)
                item['sw_outer_ring'] = re.sub(r'\s+', '', material.upper()).replace('INCONEL', 'INCONEL ').replace('ALLOY', 'ALLOY ')
        if not item.get('sw_inner_ring') and re.search(r'\bSS\s+INNER\b', upper):
            item['sw_inner_ring'] = 'SS'
        if not item.get('sw_outer_ring') and re.search(r'\bSS\s+INNER\s+AND\s+OUTER\s+CENTERING\s+RING\b', upper):
            item['sw_outer_ring'] = 'SS'

    if item.get('gasket_type') in (
            'LENS', 'MANHOLE', 'ENVELOPE', 'CMG', 'METAL_CLAD', 'SOLID_METAL',
            'LIP_SEAL', 'DIAPHRAGM', 'EYELET', 'SHEET', 'PLUG_GASKET'):
        _enrich_specialty(item, desc, upper)

    if item.get('gasket_type') in ('SOFT_CUT', 'SHEET_GASKET', 'CORRUGATED') and not item.get('moc'):
        material = _material_from_text(desc)
        if material:
            item['moc'] = material
    # Kroll & Ziller spacer gaskets: GGPL quotes 4.5MM THK and FF
    if 'KROLLER & ZILLER' in str(item.get('moc') or ''):
        item.setdefault('thickness_mm', 4.5)
        if not item.get('face_type'):
            item['face_type'] = 'FF'

    if not item.get('size') and item.get('size_type') != 'OD_ID':
        size_match = re.search(r'\bSIZE\s+IN\s+INCH\s*:?\s*(\d+(?:\.\d+)?)\s*"?', upper)
        if size_match:
            item['size'] = _size_from_text(size_match.group(1))
            item['size_type'] = 'NPS'
        else:
            nb_match = re.match(r'\s*(\d+(?:\.\d+)?)\s*MM\b', upper)
            if nb_match:
                item['size'] = f'{nb_match.group(1)}MM'
                item['size_type'] = 'NB'
            else:
                trailing_nb_match = re.search(r'\((\d+(?:\.\d+)?)\s*MM\)\s*$', upper)
                if trailing_nb_match:
                    item['size'] = f'{trailing_nb_match.group(1)}MM'
                    item['size_type'] = 'NB'

    # Universal fallback for all gasket types: "NPS 1", "1 1/2\"", "DN 25",
    # "100 NB", "18”" etc. anywhere in the description.
    if not item.get('size') and item.get('size_type') != 'OD_ID':
        size = _extract_first_size(upper)
        if size:
            item['size'] = size
            item['size_type'] = 'NB' if re.search(r'DN|NB|MM', size.upper()) else 'NPS'

    return item


def _description_from_section_row(row: tuple, col_map: dict) -> str | None:
    desc = _cell_str(row, col_map.get('description'))
    technical_desc = _cell_str(row, col_map.get('technical_description'))
    gasket_form = _cell_str(row, col_map.get('gasket_form'))
    dimension = _cell_str(row, col_map.get('dimension'))

    parts = []
    for value in (desc, gasket_form, dimension, technical_desc):
        if value and value not in parts:
            parts.append(value)

    combined = ' '.join(parts)
    if not combined or not _looks_like_gasket(combined):
        return None

    size = _cell_str(row, col_map.get('size_inch'))
    rating = _cell_str(row, col_map.get('rating'))
    thk = _cell_str(row, col_map.get('thickness'))
    od = _cell_str(row, col_map.get('od_mm'))
    id_ = _cell_str(row, col_map.get('id_mm'))

    if size:
        size_rating = size
        if rating:
            size_rating += f' X {rating}'
        if thk:
            size_rating += f' X {thk}MM THK'
        parts.append(size_rating)
    elif od and id_:
        dim = f'OD {od}MM X ID {id_}MM'
        if thk:
            dim += f' X {thk}MM THK'
        parts.append(dim)
    elif thk:
        _append_field(parts, 'THK', thk, 'MM')

    moc_val = _cell_str(row, col_map.get('moc'))
    desc_type = _infer_gasket_type(combined)
    moc_is_technical = bool(moc_val and _looks_like_gasket(moc_val))
    # For rich SPW description rows, generic MATERIAL columns such as
    # LTCS/INCO are pipe/spec context, not gasket MOC. Keep technical
    # MOC cells, because some sheets store the entire SPW construction
    # in MOC and only a short label in DESCRIPTION.
    if moc_val and moc_val.upper() != combined.upper() and (desc_type != 'SPIRAL_WOUND' or moc_is_technical):
        parts.append(f'MOC: {moc_val}')

    _append_field(parts, 'INNER RING WIDTH', _cell_str(row, col_map.get('inner_ring_width')), 'MM')
    _append_field(parts, 'OUTER RING WIDTH', _cell_str(row, col_map.get('outer_ring_width')), 'MM')
    _append_field(parts, 'REMARKS', _cell_str(row, col_map.get('remarks')))

    return ' '.join(parts)


def _parse_description_sections(ws) -> list[dict]:
    all_rows = worksheet_rows_with_merged_values(ws)
    sections = _detect_description_sections(all_rows)
    if not sections:
        return []

    items = []
    for _, col_map, data_rows in sections:
        qty_col = col_map.get('quantity')
        uom_col = col_map.get('uom')
        line_col = col_map.get('line_no')
        if line_col is None:
            line_col = _infer_line_no_col(data_rows, col_map.get('description'))

        filldown_values: dict[str, str] = {}
        for row in data_rows:
            row = _with_filldown_values(row, col_map, filldown_values)
            desc = _description_from_section_row(row, col_map)
            if not desc:
                continue
            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            if qty_col is not None and qty is None and line_no is None:
                continue
            item = {
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'raw_description': desc,
                'quantity': qty,
                'uom': uom,
                'gasket_type': _infer_gasket_type(desc) or 'SOFT_CUT',
            }
            size = _cell_str(row, col_map.get('size_inch'))
            rating = _cell_str(row, col_map.get('rating'))
            thk = _cell_str(row, col_map.get('thickness'))
            moc = _cell_str(row, col_map.get('moc'))
            od = _cell_str(row, col_map.get('od_mm'))
            id_ = _cell_str(row, col_map.get('id_mm'))
            dimension = _cell_str(row, col_map.get('dimension'))
            if size:
                item['size'] = _size_from_text(size)
                item['size_type'] = 'NPS'
            if rating:
                item['rating'] = _rating_from_text(rating)
            if thk:
                item['thickness_mm'] = _float_from_text(thk)
            if moc and item['gasket_type'] != 'SPIRAL_WOUND':
                item['moc'] = moc
            if od and id_:
                item['size_type'] = 'OD_ID'
                item['od_mm'] = _float_from_text(od)
                item['id_mm'] = _float_from_text(id_)
            if dimension:
                if not item.get('size'):
                    dim_size = _extract_first_size(dimension)
                    if dim_size:
                        item['size'] = dim_size
                        item['size_type'] = 'NPS'
                if not item.get('rating'):
                    item['rating'] = _rating_from_text(dimension)
                if not item.get('thickness_mm'):
                    thk_match = re.search(r'(?<!/)(\d+(?:\.\d+)?)\s*(?:MM)?\s*(?:THK|THICK)', dimension, re.IGNORECASE)
                    if thk_match:
                        item['thickness_mm'] = float(thk_match.group(1))
            items.append(_enrich_from_description(item))

    return items


def _with_filldown_values(row: tuple, col_map: dict, filldown_values: dict[str, str]) -> tuple:
    """Fill repeated RFQ context columns that customers leave blank on following rows."""
    mutable = list(row)
    for col_type in ('moc', 'gasket_form'):
        col_idx = col_map.get(col_type)
        if col_idx is None or col_idx >= len(mutable):
            continue
        value = _cell_str(tuple(mutable), col_idx)
        if value:
            filldown_values[col_type] = value
        elif col_type in filldown_values:
            mutable[col_idx] = filldown_values[col_type]
    return tuple(mutable)


def _infer_line_no_col(data_rows: list[tuple], desc_col: int | None) -> int | None:
    """Infer a serial-number column immediately to the left of Description."""
    if desc_col is None or desc_col <= 0:
        return None
    best_col = None
    best_count = 0
    for col_idx in range(desc_col):
        count = 0
        for row in data_rows[:25]:
            if _cell_float(row, col_idx) is not None:
                count += 1
        if count > best_count:
            best_col = col_idx
            best_count = count
    return best_col if best_count >= 1 else None


def _detect_structured_sections(all_rows: list[tuple]) -> list[tuple]:
    """Find all structured-format header blocks in a sheet.

    Returns a list of (header_row_idx_0based, col_map, data_rows) tuples where
    col_map maps field names to 0-based column indices and data_rows are the
    value tuples between this header and the next.
    """
    sections = []
    current_header_idx = None
    current_col_map = None
    current_data: list[tuple] = []

    def _is_structured_header(col_map: dict) -> bool:
        has_material = 'material' in col_map
        has_dn_cls = 'dn_size' in col_map and 'class_rating' in col_map
        has_od_id = 'od_mm' in col_map and 'id_mm' in col_map
        has_size_rating = 'size_inch' in col_map and 'rating' in col_map
        return has_material and (has_dn_cls or has_od_id or has_size_rating)

    for row_idx, row in enumerate(all_rows):
        # Try to interpret this row as a header
        col_map = {}
        for col_idx, cell in enumerate(row):
            norm = _norm_header_cell(cell)
            if not norm:
                continue
            col_type = _classify_structured_col(norm)
            if col_type and col_type not in col_map:
                col_map[col_type] = col_idx

        if _is_structured_header(col_map):
            # Save previous section if any
            if current_col_map is not None and current_data:
                sections.append((current_header_idx, current_col_map, current_data))
            current_header_idx = row_idx
            current_col_map = col_map
            current_data = []
        elif current_col_map is not None:
            # Accumulate data rows (skip fully-empty rows)
            if any(c is not None for c in row):
                current_data.append(row)

    if current_col_map is not None and current_data:
        sections.append((current_header_idx, current_col_map, current_data))

    return sections


def _parse_structured_sheet(ws) -> list[dict]:
    """Parse a sheet that uses one column per field (no combined description column)."""
    all_rows = worksheet_rows_with_merged_values(ws)
    sections = _detect_structured_sections(all_rows)
    if not sections:
        return []

    items = []
    for _, col_map, data_rows in sections:
        mat_col = col_map.get('material')
        dn_col = col_map.get('dn_size')
        cls_col = col_map.get('class_rating')
        thk_col = col_map.get('thickness')
        od_col = col_map.get('od_mm')
        id_col = col_map.get('id_mm')
        size_col = col_map.get('size_inch')
        rating_col = col_map.get('rating')
        qty_col = col_map.get('quantity')
        uom_col = col_map.get('uom')
        line_col = col_map.get('line_no')

        last_material = None

        for row in data_rows:
            # Fill-down for MATERIAL column
            mat_val = _cell_str(row, mat_col) if mat_col is not None else None
            if mat_val:
                last_material = mat_val
            else:
                mat_val = last_material

            if not mat_val:
                continue

            # Build synthetic description
            if dn_col is not None and cls_col is not None:
                dn = _cell_str(row, dn_col)
                cls = _cell_str(row, cls_col)
                if not dn or not cls:
                    continue
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' {thk}MM THK' if thk else ''
                desc = f'{dn} NB {cls}#{thk_part} GASKET MOC: {mat_val}'
                fields = {
                    'size': f'{dn} NB',
                    'size_type': 'NB',
                    'rating': _rating_from_text(cls),
                    'moc': mat_val,
                    'thickness_mm': _float_from_text(thk),
                }
            elif od_col is not None and id_col is not None:
                od = _cell_str(row, od_col)
                id_ = _cell_str(row, id_col)
                if not od or not id_:
                    continue
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' X {thk}MM THK' if thk else ''
                desc = f'OD {od}MM X ID {id_}MM{thk_part} GASKET MOC: {mat_val}'
                fields = {
                    'size_type': 'OD_ID',
                    'od_mm': _float_from_text(od),
                    'id_mm': _float_from_text(id_),
                    'moc': mat_val,
                    'thickness_mm': _float_from_text(thk),
                }
            elif size_col is not None and rating_col is not None:
                size_raw = _cell_str(row, size_col)
                rating_raw = _cell_str(row, rating_col)
                if not size_raw or not rating_raw:
                    continue
                # Numeric inch value (e.g. 10 → 10", 1.5 → 1.5")
                try:
                    sv = float(size_raw)
                    size_str = f'{int(sv)}"' if sv == int(sv) else f'{sv}"'
                except ValueError:
                    size_str = size_raw  # already a formatted string
                thk = _cell_str(row, thk_col) if thk_col is not None else None
                thk_part = f' {thk}MM THK' if thk else ''
                desc = f'{size_str} {rating_raw}{thk_part} GASKET MOC: {mat_val}'
                fields = {
                    'size': size_str,
                    'size_type': 'NPS',
                    'rating': _rating_from_text(rating_raw),
                    'moc': mat_val,
                    'thickness_mm': _float_from_text(thk),
                }
            else:
                continue

            if not _looks_like_gasket(desc):
                continue

            qty = _cell_float(row, qty_col) if qty_col is not None else None
            uom = _normalize_uom(_cell_str(row, uom_col) or 'NOS')
            line_no = _cell_float(row, line_col) if line_col is not None else None
            items.append(_enrich_from_description({
                'line_no': int(line_no) if line_no else None,
                'description': desc,
                'raw_description': desc,
                'quantity': qty,
                'uom': uom,
                'gasket_type': _infer_gasket_type(desc) or 'SOFT_CUT',
                **{key: value for key, value in fields.items() if value is not None},
            }))

    return items


# Keywords to identify column types — order matters: more specific first
_HEADER_PATTERNS = {
    'description': ['description', 'dessription', 'desription', 'desc', 'notes'],
    'technical_description': ['material and dimensional standard', 'material & dimensional standard',
                              'material standard', 'dimensional standard', 'standards', 'standard'],
    'dimension':    ['dimension/ring size', 'dimesion/ring size', 'ring size', 'dimension', 'dimesion'],
    'gasket_form':  ['gasket type', 'type'],
    'quantity':    ['qty', 'quantity', 'gross total', 'balance to order', 'balance', 'required qty', 'count'],
    'uom':         ['uom', 'inv uom'],
    'line_no':     ['sl.no', 'sl no', 'sr. no', 'sr no', 'sr no.', 'sno', 'serial', 'sr. no.', 'sr no'],
    'moc':         ['moc', 'base material'],
    'size_inch':   ['nps', 'size'],
    'rating':      ['class /rating', 'class/rating', 'rating', 'class'],
    'thickness':   ['thickness', 'thk', 'thick'],
    'od_mm':       ['gasket od', '(od)', 'od (', 'od mm'],
    'id_mm':       ['gasket id', '(id)', 'id (', 'id mm'],
    'remarks':     ['remarks', 'remark'],
    'inner_ring_width': ['inner ring width'],
    'outer_ring_width': ['outer / center ring  width', 'outer / center ring width', 'outer ring width',
                         'center ring width', 'centering ring width'],
}


def _detect_header(ws) -> tuple[int, dict]:
    """Scan first 15 rows to find the header row and map column indices."""
    best_row, best_col_map, best_score = 0, {}, 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        col_map = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_lower = str(cell).lower().strip()
            for col_type, keywords in _HEADER_PATTERNS.items():
                if col_type not in col_map and any(kw in cell_lower for kw in keywords):
                    col_map[col_type] = col_idx  # 0-based
        score = len(col_map)
        if score >= 2 and 'description' in col_map and score > best_score:
            best_score, best_row, best_col_map = score, row_idx, col_map
    return best_row, best_col_map


def _cell_str(row: tuple, idx: int | None) -> str | None:
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    # Normalize Shift+Enter newlines within a cell into a single space
    return re.sub(r'[\r\n]+', ' ', str(val)).strip()


def _cell_float(row: tuple, idx: int | None) -> float | None:
    if idx is None or idx >= len(row):
        return None
    try:
        return float(row[idx])
    except (TypeError, ValueError):
        return None
    (r'SS[-\s]*316L|316L\s*SS|316LSS', 'SS316L'),
    (r'SS[-\s]*316|316\s*SS|316SS', 'SS316'),
    (r'SS[-\s]*304L|304L\s*SS|304LSS', 'SS304L'),
    (r'SS[-\s]*304|304\s*SS|304SS', 'SS304'),
